#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_questionnaire.py
─────────────────────────
Lee el Excel quincenal y emite un cuestionario de clasificación exhaustivo
en formato Markdown, listando CADA concepto único encontrado con:
  - Sección ETL actual
  - Monto proyectado (rango min-max visto)
  - Frecuencia (en cuántas quincenas apareció)
  - Quién paga (Norma / Benjamín / Ambos — según columnas del Excel)

El cuestionario se organiza por sección ETL para facilitar la revisión.
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ──────────────────────────────────────────────────────────────────────────────
# CONSTANTES — replicadas del ETL principal para coherencia
# ──────────────────────────────────────────────────────────────────────────────

EXCEL_SECTION_TO_CATEGORY: dict[str, str] = {
    "HOUSING":                "HOUSING",
    "TRANSPORTATION":         "TRANSPORTATION",
    "PERSONAL":               "SEGUROS_MEDICOS",
    "PERSONAL CARE":          "PERSONAL_CARE",
    "FOOD":                   "FOOD",
    "PETS":                   "PETS",
    "ENTERTAINMENT":          "ENTERTAINMENT",
    "LOANS":                  "LOANS",
    "SCHOOL":                 "TRANSFERENCIAS_FAMILIARES",
    "TAXES":                  "TRANSFERENCIAS_FAMILIARES",
    "SAVINGS OR INVESTMENTS": "SAVINGS",
    "GIFTS AND DONATIONS":    "GIFTS",
    "LEGAL":                  "LEGAL",
    "OTHERS":                 "OTHER",
}

SECTION_DISPLAY: dict[str, str] = {
    "HOUSING":                        "🏠 Vivienda (HOUSING)",
    "TRANSPORTATION":                 "🚗 Transporte (TRANSPORTATION)",
    "SEGUROS_MEDICOS":                "🏥 Seguros Médicos (PERSONAL)",
    "PERSONAL_CARE":                  "💆 Cuidado Personal (PERSONAL CARE)",
    "FOOD":                           "🛒 Alimentación (FOOD)",
    "PETS":                           "🐱 Mascotas (PETS)",
    "ENTERTAINMENT":                  "🎬 Entretenimiento (ENTERTAINMENT)",
    "LOANS":                          "💳 Préstamos y Tarjetas (LOANS)",
    "TRANSFERENCIAS_FAMILIARES":      "👨‍👩‍👧‍👦 Transferencias Familiares (SCHOOL/TAXES)",
    "SAVINGS":                        "💰 Ahorros (SAVINGS OR INVESTMENTS)",
    "GIFTS":                          "🎁 Regalos y Donaciones (GIFTS)",
    "LEGAL":                          "⚖️ Legal (LEGAL)",
    "OTHER":                          "📦 Otros (OTHERS)",
}

# Categorías hijas conocidas — para mostrar la asignación actual del ETL
KNOWN_CHILDREN: dict[str, list[tuple[str, str]]] = {
    "HOUSING": [
        ("HOUSING.HIPOTECA", "Hipoteca"),
        ("HOUSING.INTERNET", "Internet"),
        ("HOUSING.ELECTRICIDAD", "Electricidad"),
        ("HOUSING.AGUA", "Agua"),
        ("HOUSING.TELEFONO", "Teléfono"),
        ("HOUSING.FRACCIONAMIENTO", "Fraccionamiento"),
    ],
    "TRANSPORTATION": [
        ("TRANSPORTATION.GASOLINA", "Gasolina"),
        ("TRANSPORTATION.INSURANCE", "Seguro vehículo"),
        ("TRANSPORTATION.LICENSING", "Licencias"),
        ("TRANSPORTATION.MAINTENANCE", "Mantenimiento"),
    ],
    "SEGUROS_MEDICOS": [
        ("SEGUROS_MEDICOS.BENJI", "Seguro Benji"),
        ("SEGUROS_MEDICOS.NORMA", "Seguro Norma"),
        ("SEGUROS_MEDICOS.HIJOS", "Seguro hijos"),
        ("SEGUROS_MEDICOS.SANTI", "Seguro Santi"),
    ],
    "ENTERTAINMENT": [
        ("ENTERTAINMENT.NETFLIX", "Netflix"),
        ("ENTERTAINMENT.HBO", "HBO"),
        ("ENTERTAINMENT.PRIME", "Amazon Prime"),
        ("ENTERTAINMENT.DISNEY", "Disney+/Star"),
        ("ENTERTAINMENT.SPOTIFY", "Spotify"),
        ("ENTERTAINMENT.HAWAIANO", "Hawaiano"),
        ("ENTERTAINMENT.DIVERSION", "Diversión"),
    ],
    "FOOD": [
        ("FOOD.COMIDA", "Comida"),
        ("FOOD.DESPENSA", "Despensa"),
        ("FOOD.LIMPIEZA", "Limpieza"),
    ],
    "PETS": [
        ("PETS.COMIDA", "Comida gatas"),
        ("PETS.GROOMING", "Grooming"),
        ("PETS.VETERINARIO", "Veterinario"),
    ],
    "LOANS": [
        ("LOANS.COPPEL", "Coppel"),
        ("LOANS.LIVERPOOL", "Liverpool"),
        ("LOANS.SEARS", "Sears"),
        ("LOANS.WALMART", "Walmart"),
        ("LOANS.BANAMEX_CC", "Banamex Clásica"),
        ("LOANS.MERCADO_LIBRE", "Mercado Libre"),
        ("LOANS.MERCADO_PAGO", "Mercado Pago"),
        ("LOANS.OMAR", "Préstamo Omar"),
    ],
    "TRANSFERENCIAS_FAMILIARES": [
        ("TRANSFERENCIAS_FAMILIARES.DAVID", "David"),
        ("TRANSFERENCIAS_FAMILIARES.PAU", "Pau"),
        ("TRANSFERENCIAS_FAMILIARES.SANTIAGO", "Santi"),
        ("TRANSFERENCIAS_FAMILIARES.COCHE", "Coche"),
        ("TRANSFERENCIAS_FAMILIARES.INSCRIPCIONES", "Inscripciones"),
    ],
    "SAVINGS": [
        ("SAVINGS.EMPRESA", "Ahorro Empresa"),
        ("SAVINGS.TARJETA", "Tarjeta de ahorro"),
        ("SAVINGS.RETIREMENT", "Retiro"),
        ("SAVINGS.INVESTMENT", "Inversión"),
        ("SAVINGS.EFECTIVO", "Ahorro efectivo"),
    ],
}

MEMBERS = ["Benjamín", "Norma", "Pau", "David", "Agustín", "Santiago"]

# ──────────────────────────────────────────────────────────────────────────────
# UTILIDADES
# ──────────────────────────────────────────────────────────────────────────────

def _normalize(text: str) -> str:
    if not text:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(text))
    stripped = "".join(ch for ch in nfkd if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", stripped).strip().upper()


def _as_float(value) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        f = float(value)
        return f if f != 0 else None
    s = str(value).strip()
    if not s or s.startswith("#"):
        return None
    cleaned = re.sub(r"[^\d.\-]", "", s.replace(",", ""))
    try:
        f = float(cleaned)
        return f if f != 0 else None
    except ValueError:
        return None


def _as_str(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


# ──────────────────────────────────────────────────────────────────────────────
# ESTRUCTURAS DE DATOS
# ──────────────────────────────────────────────────────────────────────────────

@dataclass
class ConceptRecord:
    """Información agregada de un concepto a lo largo de todas las quincenas."""
    concept: str               # Texto original más frecuente
    section_code: str          # Código de sección ETL
    occurrences: int = 0       # Número de quincenas en que apareció
    amounts: list[float] = field(default_factory=list)
    norma_pays: int = 0        # Quincenas donde Norma tiene monto > 0
    benja_pays: int = 0        # Quincenas donde Benjamín tiene monto > 0
    variants: set[str] = field(default_factory=set)  # Variantes de escritura


# ──────────────────────────────────────────────────────────────────────────────
# EXTRACCIÓN DEL EXCEL
# ──────────────────────────────────────────────────────────────────────────────

def find_section_headers(ws: Worksheet) -> list[tuple[int, int, int, int, int, str]]:
    """
    Detecta headers de sección por la presencia de 'Projected Cost'.
    Retorna: [(row, name_col, projected_col, norma_col, benja_col, section_code)]
    """
    result = []
    max_row = min(ws.max_row, 65)
    max_col = min(ws.max_column, 14)
    for r in range(1, max_row + 1):
        for c in range(2, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if not isinstance(val, str) or "Projected Cost" not in val:
                continue
            name_cell = ws.cell(row=r, column=c - 1).value
            if not name_cell:
                continue
            section_norm = _normalize(str(name_cell))
            section_code = EXCEL_SECTION_TO_CATEGORY.get(section_norm)
            if section_code is None:
                continue
            result.append((r, c - 1, c, c + 1, c + 2, section_code))
    return result


_BLOCK_TERMINATORS = {"SUBTOTAL", "TOTAL", "TOTAL SUBTOTAL"}


def extract_lines_from_sheet(ws: Worksheet) -> list[tuple[str, str, float, Optional[float], Optional[float]]]:
    """
    Extrae (concept, section_code, projected, norma_amount, benja_amount)
    de todas las secciones de la hoja.
    """
    lines = []
    for row, name_col, proj_col, norma_col, benja_col, section_code in find_section_headers(ws):
        empty_streak = 0
        for r in range(row + 1, min(row + 25, ws.max_row + 1)):
            concept_raw = ws.cell(row=r, column=name_col).value
            projected = _as_float(ws.cell(row=r, column=proj_col).value)
            norma = _as_float(ws.cell(row=r, column=norma_col).value)
            benja = _as_float(ws.cell(row=r, column=benja_col).value)

            concept_str = _as_str(concept_raw) or ""
            concept_upper = _normalize(concept_str)

            if concept_upper in _BLOCK_TERMINATORS:
                break
            if not concept_str and projected is None:
                empty_streak += 1
                if empty_streak >= 3:
                    break
                continue
            empty_streak = 0

            if projected is None or projected <= 0:
                continue
            if not concept_str or concept_upper in {"OTHER"}:
                continue

            lines.append((concept_str, section_code, projected, norma, benja))
    return lines


def build_concept_registry(excel_path: Path) -> dict[str, ConceptRecord]:
    """
    Lee todas las hojas del Excel y construye un registro de conceptos únicos.
    La clave es `normalize(concept) + "|" + section_code` para distinguir
    conceptos iguales en distintas secciones (raro pero posible).
    """
    wb = load_workbook(excel_path, data_only=True, read_only=False)
    registry: dict[str, ConceptRecord] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines = extract_lines_from_sheet(ws)
        for concept, section_code, projected, norma, benja in lines:
            key = f"{_normalize(concept)}|{section_code}"
            if key not in registry:
                registry[key] = ConceptRecord(
                    concept=concept,
                    section_code=section_code,
                )
            rec = registry[key]
            rec.occurrences += 1
            rec.amounts.append(projected)
            rec.variants.add(concept.strip())
            if norma and norma > 0:
                rec.norma_pays += 1
            if benja and benja > 0:
                rec.benja_pays += 1
            # Usar la variante más larga (más descriptiva) como nombre canónico
            if len(concept) > len(rec.concept):
                rec.concept = concept

    return registry


# ──────────────────────────────────────────────────────────────────────────────
# INFERENCIA DE CATEGORÍA ETL ACTUAL
# ──────────────────────────────────────────────────────────────────────────────

def infer_current_category(concept: str, section_code: str) -> str:
    """Replica la lógica de resolve_category() del ETL para mostrar lo que haría."""
    concept_norm = _normalize(concept)
    children = KNOWN_CHILDREN.get(section_code, [])
    for code, display in children:
        leaf = code.rsplit(".", 1)[-1]
        leaf_display = _normalize(display)
        if leaf in concept_norm or leaf_display in concept_norm:
            return code

    # Casos especiales
    special_map = {
        "BENJI":  f"SEGUROS_MEDICOS.BENJI"  if section_code == "SEGUROS_MEDICOS" else None,
        "NORMA":  f"SEGUROS_MEDICOS.NORMA"  if section_code == "SEGUROS_MEDICOS" else None,
        "SANTI":  f"SEGUROS_MEDICOS.SANTI"  if section_code == "SEGUROS_MEDICOS" else None,
        "DAVID":  f"TRANSFERENCIAS_FAMILIARES.DAVID" if section_code == "TRANSFERENCIAS_FAMILIARES" else None,
        "PAU":    f"TRANSFERENCIAS_FAMILIARES.PAU"   if section_code == "TRANSFERENCIAS_FAMILIARES" else None,
        "COCHE":  f"TRANSFERENCIAS_FAMILIARES.COCHE" if section_code == "TRANSFERENCIAS_FAMILIARES" else None,
    }
    for needle, code in special_map.items():
        if code and needle in concept_norm:
            return code

    if "PAU" in concept_norm and "DAVID" in concept_norm and section_code == "SEGUROS_MEDICOS":
        return "SEGUROS_MEDICOS.HIJOS"

    return section_code  # fallback: raíz


def infer_beneficiaries(concept: str, section_code: str) -> str:
    """Replica heurística de resolve_beneficiaries()."""
    concept_norm = _normalize(concept)
    ALIASES = {
        "Benjamín": ["BENJI", "BENJAMIN", "BENJAMÍN"],
        "Norma":    ["NORMA"],
        "Pau":      ["PAU", "PAULINA"],
        "David":    ["DAVID", "DAV", "DAVE"],
        "Agustín":  ["AGUS", "AGUSTIN", "AGUSTÍN"],
        "Santiago": ["SANTI", "SANTIAGO"],
    }
    hits = []
    for name, aliases in ALIASES.items():
        for alias in aliases:
            if alias in concept_norm:
                if name not in hits:
                    hits.append(name)
                break
    return ", ".join(hits) if hits else "Norma, Benjamín (default)"


def infer_payment_method(concept: str, norma_pays: int, benja_pays: int) -> str:
    """Replica heurística de resolve_payment_method()."""
    concept_norm = _normalize(concept)
    KEYWORDS = [
        ("BANAMEX CLASICA", "Banamex Clásica"),
        ("MERCADO LIBRE", "Mercado Libre"),
        ("MERCADO PAGO", "Mercado Pago"),
        ("COPPEL", "Coppel"),
        ("LIVERPOOL", "Liverpool"),
        ("SEARS", "Sears"),
        ("WALMART", "Walmart"),
        ("KLAR", "Klar"),
        ("BBVA", "BBVA"),
    ]
    for needle, display in KEYWORDS:
        if needle in concept_norm:
            return display
    if norma_pays > benja_pays:
        return "BBVA (Norma)"
    if benja_pays > norma_pays:
        return "Efectivo (Benji)"
    return "Efectivo"


# ──────────────────────────────────────────────────────────────────────────────
# GENERACIÓN DEL CUESTIONARIO MARKDOWN
# ──────────────────────────────────────────────────────────────────────────────

def format_mxn(amount: float) -> str:
    return f"${amount:,.0f}"


def generate_questionnaire(registry: dict[str, ConceptRecord]) -> str:
    # Agrupar por sección
    by_section: dict[str, list[ConceptRecord]] = defaultdict(list)
    for rec in registry.values():
        by_section[rec.section_code].append(rec)

    # Ordenar secciones por orden naturalal
    section_order = list(EXCEL_SECTION_TO_CATEGORY.values())
    # Deduplicar manteniendo orden
    seen = set()
    section_order_unique = []
    for s in section_order:
        if s not in seen:
            section_order_unique.append(s)
            seen.add(s)

    lines = []

    lines.append("# 📋 Cuestionario de Clasificación de Gastos")
    lines.append("")
    lines.append("> **Instrucciones**: Revisa cada concepto. Si la clasificación ETL es correcta, no hagas nada.")
    lines.append("> Si es incorrecta, completa las columnas `Cat. Correcta`, `Beneficiarios Correctos` y `Pago Correcto`.")
    lines.append("> Usa `✅` para confirmar o escribe el valor correcto.")
    lines.append("> Los valores que dejes en blanco se tomarán como **correctos tal cual**.")
    lines.append("")
    lines.append("## Miembros del hogar")
    lines.append("| Clave | Nombre | Rol |")
    lines.append("|-------|--------|-----|")
    lines.append("| `ben` | Benjamín | Adulto-pagador |")
    lines.append("| `nor` | Norma | Adulto-pagadora |")
    lines.append("| `pau` | Pau | Dependiente |")
    lines.append("| `dav` | David | Dependiente |")
    lines.append("| `agu` | Agustín | Dependiente |")
    lines.append("| `san` | Santiago | Dependiente |")
    lines.append("| `todos` | Todos los anteriores | — |")
    lines.append("")
    lines.append("## Categorías disponibles")
    lines.append("*(Si necesitas crear una nueva, escríbela como `NUEVA: nombre`)*")
    lines.append("")
    lines.append("| Código | Descripción |")
    lines.append("|--------|-------------|")
    all_cats = [
        ("HOUSING", "Vivienda (raíz)"),
        ("HOUSING.HIPOTECA", "Hipoteca"),
        ("HOUSING.INTERNET", "Internet"),
        ("HOUSING.ELECTRICIDAD", "Electricidad"),
        ("HOUSING.AGUA", "Agua"),
        ("HOUSING.TELEFONO", "Teléfono"),
        ("HOUSING.FRACCIONAMIENTO", "Fraccionamiento"),
        ("TRANSPORTATION", "Transporte (raíz)"),
        ("TRANSPORTATION.GASOLINA", "Gasolina"),
        ("TRANSPORTATION.INSURANCE", "Seguro vehículo"),
        ("TRANSPORTATION.MAINTENANCE", "Mantenimiento"),
        ("SEGUROS_MEDICOS", "Seguros Médicos (raíz)"),
        ("SEGUROS_MEDICOS.BENJI", "Seguro Benji"),
        ("SEGUROS_MEDICOS.NORMA", "Seguro Norma"),
        ("SEGUROS_MEDICOS.HIJOS", "Seguro hijos (Pau+David+Agus)"),
        ("SEGUROS_MEDICOS.SANTI", "Seguro Santi"),
        ("FOOD", "Alimentación (raíz)"),
        ("FOOD.COMIDA", "Comida"),
        ("FOOD.DESPENSA", "Despensa"),
        ("FOOD.LIMPIEZA", "Limpieza"),
        ("PETS", "Mascotas (raíz)"),
        ("PETS.COMIDA", "Comida gatas"),
        ("PETS.GROOMING", "Grooming"),
        ("PETS.VETERINARIO", "Veterinario"),
        ("ENTERTAINMENT", "Entretenimiento (raíz)"),
        ("ENTERTAINMENT.NETFLIX", "Netflix"),
        ("ENTERTAINMENT.HBO", "HBO"),
        ("ENTERTAINMENT.PRIME", "Amazon Prime"),
        ("ENTERTAINMENT.DISNEY", "Disney+"),
        ("ENTERTAINMENT.SPOTIFY", "Spotify"),
        ("ENTERTAINMENT.HAWAIANO", "Hawaiano"),
        ("ENTERTAINMENT.DIVERSION", "Diversión"),
        ("LOANS", "Préstamos/Tarjetas (raíz)"),
        ("LOANS.COPPEL", "Coppel"),
        ("LOANS.LIVERPOOL", "Liverpool"),
        ("LOANS.SEARS", "Sears"),
        ("LOANS.WALMART", "Walmart"),
        ("LOANS.BANAMEX_CC", "Banamex Clásica"),
        ("LOANS.MERCADO_LIBRE", "Mercado Libre"),
        ("LOANS.MERCADO_PAGO", "Mercado Pago"),
        ("LOANS.OMAR", "Préstamo Omar"),
        ("TRANSFERENCIAS_FAMILIARES", "Transferencias Familiares (raíz)"),
        ("TRANSFERENCIAS_FAMILIARES.DAVID", "Transferencia David"),
        ("TRANSFERENCIAS_FAMILIARES.PAU", "Transferencia Pau"),
        ("TRANSFERENCIAS_FAMILIARES.SANTIAGO", "Transferencia Santi"),
        ("TRANSFERENCIAS_FAMILIARES.COCHE", "Coche"),
        ("TRANSFERENCIAS_FAMILIARES.INSCRIPCIONES", "Inscripciones"),
        ("SAVINGS", "Ahorros (raíz)"),
        ("SAVINGS.EMPRESA", "Ahorro Empresa"),
        ("SAVINGS.TARJETA", "Tarjeta de ahorro"),
        ("SAVINGS.RETIREMENT", "Retiro"),
        ("SAVINGS.INVESTMENT", "Inversión"),
        ("SAVINGS.EFECTIVO", "Ahorro efectivo"),
        ("PERSONAL_CARE", "Cuidado Personal"),
        ("GIFTS", "Regalos y Donaciones"),
        ("LEGAL", "Legal"),
        ("OTHER", "Otros"),
        ("SERVICIOS_EXTERNOS.ARACELI", "Araceli (empleada del hogar)"),
        ("SERVICIOS_EXTERNOS.PSICOLOGA", "Psicóloga"),
    ]
    for code, desc in all_cats:
        lines.append(f"| `{code}` | {desc} |")

    lines.append("")
    lines.append("---")
    lines.append("")

    total_concepts = 0

    for section_code in section_order_unique:
        records = by_section.get(section_code, [])
        if not records:
            continue

        # Ordenar por frecuencia descendente, luego alfabético
        records.sort(key=lambda r: (-r.occurrences, r.concept.lower()))

        section_display = SECTION_DISPLAY.get(section_code, section_code)
        lines.append(f"## {section_display}")
        lines.append("")

        col_headers = [
            "N°",
            "Concepto",
            "Variantes encontradas",
            "Quincenas",
            "Monto (min–max)",
            "Quién paga",
            "**Cat. ETL actual**",
            "**Benef. ETL actual**",
            "**Pago ETL actual**",
            "✏️ Cat. Correcta",
            "✏️ Beneficiarios Correctos",
            "✏️ Metodo Pago Correcto",
        ]
        lines.append("| " + " | ".join(col_headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(col_headers)) + " |")

        for idx, rec in enumerate(records, 1):
            total_concepts += 1
            min_amt = min(rec.amounts)
            max_amt = max(rec.amounts)
            amt_str = format_mxn(min_amt) if min_amt == max_amt else f"{format_mxn(min_amt)}–{format_mxn(max_amt)}"

            # Quién paga
            if rec.norma_pays > 0 and rec.benja_pays > 0:
                payer_info = "Ambos"
            elif rec.norma_pays > 0:
                payer_info = "Norma"
            elif rec.benja_pays > 0:
                payer_info = "Benji"
            else:
                payer_info = "Sin dato"

            # Variantes — mostrar solo si hay diferencias reales
            variants_clean = sorted(rec.variants - {rec.concept})
            variants_str = ", ".join(f"`{v}`" for v in variants_clean[:3]) if variants_clean else "—"
            if len(variants_clean) > 3:
                variants_str += f" _(+{len(variants_clean)-3})_"

            # Inferir estado actual del ETL
            current_cat = infer_current_category(rec.concept, rec.section_code)
            current_benef = infer_beneficiaries(rec.concept, rec.section_code)
            current_pay = infer_payment_method(rec.concept, rec.norma_pays, rec.benja_pays)

            row = [
                str(total_concepts),
                f"**{rec.concept}**",
                variants_str,
                str(rec.occurrences),
                amt_str,
                payer_info,
                f"`{current_cat}`",
                current_benef,
                current_pay,
                "",  # ✏️ Cat. Correcta — usuario rellena
                "",  # ✏️ Beneficiarios Correctos
                "",  # ✏️ Metodo Pago Correcto
            ]
            lines.append("| " + " | ".join(row) + " |")

        lines.append("")

    # Resumen final
    lines.append("---")
    lines.append("")
    lines.append(f"## Resumen")
    lines.append(f"")
    lines.append(f"- **Total conceptos únicos encontrados**: {total_concepts}")
    lines.append(f"- **Secciones analizadas**: {len([s for s in section_order_unique if by_section.get(s)])}")
    lines.append(f"- **Hojas procesadas**: {sum(1 for r in registry.values() for _ in [r])}")

    return "\n".join(lines)


# ──────────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Genera cuestionario de clasificación de gastos desde el Excel quincenal."
    )
    parser.add_argument(
        "--excel", type=Path,
        default=Path("Copy of presupuesto 2.5.xlsx"),
        help="Ruta al Excel de origen."
    )
    parser.add_argument(
        "--output", type=Path,
        default=Path("clasificacion_cuestionario.md"),
        help="Ruta del Markdown de salida."
    )
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"ERROR: No se encontró el Excel: {args.excel}", file=sys.stderr)
        sys.exit(2)

    print(f"[INFO] Leyendo Excel: {args.excel}")
    registry = build_concept_registry(args.excel)
    print(f"[OK] {len(registry)} conceptos unicos encontrados")

    print(f"[INFO] Generando cuestionario...")
    md = generate_questionnaire(registry)

    args.output.write_text(md, encoding="utf-8")
    print(f"[OK] Cuestionario guardado en: {args.output}")
    print(f"     Abrelo con cualquier editor Markdown para rellenarlo.")


if __name__ == "__main__":
    main()
