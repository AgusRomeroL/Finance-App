#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
 excel_to_room_etl.py  —  Puente de datos: Excel quincenal → Room SQLite
================================================================================

Lee el Excel de presupuesto quincenal (hojas manuales con layout
variable, typos, y formatos de fecha inconsistentes) y emite un archivo
``budget_database.db`` cuyo esquema es bit-compatible con el generado por Room
para ``mx.budget.data.local.BudgetDatabase`` (version = 1).

Entidades pobladas
------------------
  1. household                 — singleton del hogar
  2. member                    — Benjamín, Norma, Pau, David, Agustín, Santi, …
  3. category                  — árbol de categorías jerárquico
  4. payment_method            — Banamex, BBVA, Mercado Pago, Efectivo, …
  5. quincena                  — 33 períodos quincenales (Ene 2025 → Jun 2026)
  6. income_source             — ingresos quincenales de Benjamín y Norma
  7. expense                   — una fila por cada línea presupuestada > 0
  8. expense_attribution       — reparto BENEFICIARY / PAYER en basis points
  9. installment_plan          — planes MSI reales (Mercado Libre, Buró)
 10. recurrence_template       — plantillas de obligaciones fijas recurrentes
 11. loan                      — préstamo a Jaudiel
 12. savings_goal              — Ahorro Empresa como meta seed

Robustez frente a datos sucios
------------------------------
  * Nombres de hoja con y sin año, con typos ("febrereo"), con separadores
    inconsistentes ("1al 15"), fechas imposibles ("31 septiembre") → parser
    tolerante con regex + fallback al calendario real via ``monthrange``.
  * Celdas con ``None``, ``""``, ``0`` o fórmulas rotas (#REF!) → helper
    ``_as_float`` que las colapsa a ``None`` sin explotar.
  * Posiciones variables de los bloques de categoría entre hojas → escaneo por
    marcador textual ("Projected Cost") en lugar de coordenadas hardcodeadas.
  * Año inferido cuando el nombre no lo indica: las primeras 21 hojas son 2025,
    el resto (que sí lo llevan) se usa como escrito.
  * IDs determinísticos (``uuid5`` con namespace fijo) para permitir re-ejecutar
    el script de forma idempotente.

Uso
---
    $ pip install pandas openpyxl
    $ python excel_to_room_etl.py \
        --excel "Copy of presupuesto 2.5.xlsx" \
        --output budget_database.db [--verbose]

Integración con la app Android
------------------------------
  1. Copiar ``budget_database.db`` a ``app/src/main/assets/``.
  2. En ``BudgetDatabase.buildDatabase``, reemplazar el builder por:

       Room.databaseBuilder(ctx, BudgetDatabase::class.java, "budget.db")
           .createFromAsset("budget_database.db")
           .fallbackToDestructiveMigration()
           .build()

  3. NOTA CRÍTICA: Room valida la coherencia esquemática contra la tabla
     interna ``room_master_table`` y su ``identity_hash``. Este script emite
     dicha tabla con un hash placeholder. Para producción, extraer el hash
     real de ``app/schemas/…/1.json`` generado por KSP en tiempo de compilación
     y actualizar ``ROOM_IDENTITY_HASH`` en la constante de abajo.
"""

from __future__ import annotations

import argparse
import calendar
import json
import logging
import re
import sqlite3
import sys
import unicodedata
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Iterable, Optional
from zoneinfo import ZoneInfo

import pandas as pd  # noqa: F401  — se usa para analítica de los conteos finales.
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ══════════════════════════════════════════════════════════════════════════════
#  BLOQUE 1 · CONFIGURACIÓN Y CONSTANTES SEMÁNTICAS
# ══════════════════════════════════════════════════════════════════════════════

LOG = logging.getLogger("etl")

# Namespace UUID arbitrario pero estable. Permite que cada corrida del script
# produzca los mismos IDs, de modo que la DB sea re-generable sin divergencia.
_UUID_NAMESPACE = uuid.UUID("b0d9e7a0-1234-5678-9abc-def012345678")


def did(key: str) -> str:
    """UUID determinístico — `uuid5` sobre el namespace fijo."""
    return str(uuid.uuid5(_UUID_NAMESPACE, key))


HOUSEHOLD_ID = did("household:default")
HOUSEHOLD_NAME = "Familia"
HOUSEHOLD_TZ = "America/Mexico_City"

# Placeholder: debe reemplazarse con el hash real del schema JSON emitido por
ROOM_IDENTITY_HASH = "9b889298fe2edb27e8865dc8239ef354"

# Epoch millis fijo para las filas "nacidas" por el ETL. Se elige el instante
# en el que se generó la base de datos, en UTC.
NOW_EPOCH_MS = int(datetime.now(tz=timezone.utc).timestamp() * 1000)

# "Hoy" en la zona horaria del hogar. Gobierna los status derivados de la
# fecha de generación: quincenas CLOSED/ACTIVE/PROVISIONED, gastos
# POSTED/PLANNED e ingresos POSTED/PLANNED. Se usa la zona del hogar (no UTC)
# porque el corte quincenal es un concepto local: a las 19:00 de CDMX ya es
# "mañana" en UTC y sin esto la quincena activa podría cerrarse antes de tiempo.
TODAY = datetime.now(ZoneInfo(HOUSEHOLD_TZ)).date()


# ── 1.1 Miembros del hogar ────────────────────────────────────────────────────

@dataclass(frozen=True)
class MemberSeed:
    key: str
    display_name: str
    role: str                 # PAYER_ADULT | BENEFICIARY_DEPENDENT | EXTERNAL_*
    aliases: tuple[str, ...]  # usados para fuzzy-match en atribución
    default_income: Optional[float] = None

    @property
    def id(self) -> str:
        return did(f"member:{self.key}")


MEMBERS: tuple[MemberSeed, ...] = (
    MemberSeed("benjamin", "Benjamín",  "PAYER_ADULT",
               ("benji", "benjamin", "benjamín"), 45_000.0),
    MemberSeed("norma",    "Norma",     "PAYER_ADULT",
               ("norma",), 60_000.0),
    # Pau se llama "Normita" en la app (pedido de Agustín 2026-07-07); el key
    # y los alias conservan "pau" porque así aparece en el Excel.
    MemberSeed("pau",      "Normita",   "BENEFICIARY_DEPENDENT",
               ("pau", "paulina", "pau,", "pau ", "normita")),
    MemberSeed("david",    "David",     "BENEFICIARY_DEPENDENT",
               ("david", "dav", "dave")),
    MemberSeed("agustin",  "Agustín",   "BENEFICIARY_DEPENDENT",
               ("agus", "agustin", "agustín")),
    MemberSeed("santiago", "Santiago",  "BENEFICIARY_DEPENDENT",
               ("santi", "santiago")),
    MemberSeed("omar",     "Omar",      "EXTERNAL_CREDITOR",
               ("omar", "prestamo omar", "préstamo omar")),
    MemberSeed("jaudiel",  "Jaudiel",   "EXTERNAL_DEBTOR",
               ("jaudiel",)),
    MemberSeed("araceli",  "Araceli",   "EXTERNAL_SERVICE",
               ("araceli",)),
    # Deudor de la cuenta por cobrar "Deben $3,600 — Medicina + rotafolio"
    # (notas del Excel oct-dic 2025). Norma sabe quién es; renombrar en la app.
    MemberSeed("deudor_pendiente", "Por identificar", "EXTERNAL_DEBTOR",
               ("deudor",)),
)

MEMBERS_BY_KEY: dict[str, MemberSeed] = {m.key: m for m in MEMBERS}


# ── 1.2 Métodos de pago (wallets) ─────────────────────────────────────────────

@dataclass(frozen=True)
class PaymentMethodSeed:
    key: str
    display_name: str
    kind: str
    issuer: Optional[str] = None
    owner_key: Optional[str] = None

    @property
    def id(self) -> str:
        return did(f"payment_method:{self.key}")


PAYMENT_METHODS: tuple[PaymentMethodSeed, ...] = (
    PaymentMethodSeed("banamex_deb",    "Banamex Débito",      "DEBIT_ACCOUNT",           "Citibanamex", "norma"),
    PaymentMethodSeed("bbva",           "BBVA",                "DEBIT_ACCOUNT",           "BBVA México", "norma"),
    PaymentMethodSeed("banamex_cc",     "Banamex Clásica",     "CREDIT_CARD",             "Citibanamex", "norma"),
    PaymentMethodSeed("mercado_pago",   "Mercado Pago",        "DIGITAL_WALLET",          "Mercado Pago", None),
    PaymentMethodSeed("mercado_libre",  "Mercado Libre BNPL",  "BNPL_INSTALLMENT",        "Mercado Libre", None),
    PaymentMethodSeed("coppel",         "Coppel",              "DEPARTMENT_STORE_CARD",   "Coppel", None),
    PaymentMethodSeed("liverpool",      "Liverpool",           "DEPARTMENT_STORE_CARD",   "Liverpool", None),
    PaymentMethodSeed("sears",          "Sears",               "DEPARTMENT_STORE_CARD",   "Sears", None),
    PaymentMethodSeed("walmart",        "Walmart",             "DEPARTMENT_STORE_CARD",   "Walmart", None),
    PaymentMethodSeed("klar",           "Klar",                "DIGITAL_WALLET",          "Klar", None),
    PaymentMethodSeed("efectivo",       "Efectivo",            "CASH",                    None, None),
    PaymentMethodSeed("ahorro_empresa", "Ahorro Empresa",      "EMPLOYER_SAVINGS_FUND",   None, "benjamin"),
)

PAYMENT_METHODS_BY_KEY: dict[str, PaymentMethodSeed] = {pm.key: pm for pm in PAYMENT_METHODS}


# ── 1.3 Categorías (árbol jerárquico) ─────────────────────────────────────────

@dataclass(frozen=True)
class CategorySeed:
    code: str                    # canónico: HOUSING | HOUSING.TELEFONO | …
    display_name: str
    parent_code: Optional[str]
    kind: str                    # EXPENSE_FIXED | EXPENSE_VARIABLE | SAVINGS | …
    budget_default: Optional[float] = None
    sort_order: int = 0

    @property
    def id(self) -> str:
        return did(f"category:{self.code}")


# NOTA: las hojas del Excel contienen al menos 14 bloques distintos
# (HOUSING, TRANSPORTATION, PERSONAL, FOOD, PETS, ENTERTAINMENT, LOANS,
#  SCHOOL, TAXES, SAVINGS OR INVESTMENTS, GIFTS AND DONATIONS, LEGAL,
#  PERSONAL CARE, OTHERS). La especificación mapea:
#     PERSONAL  → SEGUROS_MEDICOS
#     TAXES     → TRANSFERENCIAS_FAMILIARES
#     SCHOOL    → TRANSFERENCIAS_FAMILIARES (mismo destino que TAXES)
# Se respetan esas equivalencias en el loader de secciones.
CATEGORIES: tuple[CategorySeed, ...] = (
    # raíces
    CategorySeed("HOUSING",               "Vivienda",                 None, "EXPENSE_FIXED",     None,   10),
    CategorySeed("TRANSPORTATION",        "Transporte",               None, "EXPENSE_VARIABLE",  None,   20),
    CategorySeed("SEGUROS_MEDICOS",       "Seguros Médicos",          None, "EXPENSE_FIXED",     None,   30),
    CategorySeed("FOOD",                  "Alimentación",             None, "EXPENSE_VARIABLE",  None,   40),
    CategorySeed("PETS",                  "Mascotas",                 None, "EXPENSE_VARIABLE",  None,   50),
    CategorySeed("ENTERTAINMENT",         "Entretenimiento",          None, "EXPENSE_VARIABLE",  None,   60),
    CategorySeed("LOANS",                 "Tarjetas y préstamos",     None, "EXPENSE_INSTALLMENT", None, 70),
    CategorySeed("TRANSFERENCIAS_FAMILIARES", "Transferencias familiares", None, "TRANSFER_INTRA_HOUSEHOLD", None, 80),
    CategorySeed("ESCUELA",               "Escuela / Colegiaturas",   None, "EXPENSE_FIXED",     None,   85),
    CategorySeed("SAVINGS",               "Ahorros e inversiones",    None, "SAVINGS",           None,   90),
    CategorySeed("GIFTS",                 "Regalos y donaciones",     None, "EXPENSE_VARIABLE",  None,  100),
    CategorySeed("LEGAL",                 "Legal",                    None, "EXPENSE_VARIABLE",  None,  110),
    CategorySeed("PERSONAL_CARE",         "Cuidado personal",         None, "EXPENSE_VARIABLE",  None,  120),
    CategorySeed("SERVICIOS_EXTERNOS",    "Servicios externos",       None, "EXPENSE_VARIABLE",  None,  130),
    CategorySeed("OTHER",                 "Otros",                    None, "EXPENSE_VARIABLE",  None,  200),
    CategorySeed("INGRESOS",              "Ingresos",                 None, "INCOME",            None,   5),
    CategorySeed("PRESTAMOS_OTORGADOS",   "Préstamos otorgados",      None, "LOAN_RECEIVABLE",   None,  140),

    # hojas hijas más usadas (código jerárquico con punto)
    CategorySeed("HOUSING.HIPOTECA",         "Hipoteca",         "HOUSING",        "EXPENSE_FIXED",    3750.0),
    CategorySeed("HOUSING.INTERNET",         "Internet",         "HOUSING",        "EXPENSE_FIXED",     899.0),
    CategorySeed("HOUSING.ELECTRICIDAD",     "Electricidad",     "HOUSING",        "EXPENSE_VARIABLE", 1000.0),
    CategorySeed("HOUSING.AGUA",             "Agua",             "HOUSING",        "EXPENSE_VARIABLE",  None),
    CategorySeed("HOUSING.TELEFONO",         "Teléfono",         "HOUSING",        "EXPENSE_FIXED",     None),
    CategorySeed("HOUSING.FRACCIONAMIENTO",  "Fraccionamiento",  "HOUSING",        "EXPENSE_FIXED",     None),
    CategorySeed("HOUSING.MUEBLES",          "Muebles",          "HOUSING",        "EXPENSE_VARIABLE",  None),

    CategorySeed("TRANSPORTATION.GASOLINA",  "Gasolina",         "TRANSPORTATION", "EXPENSE_VARIABLE", 2600.0),
    CategorySeed("TRANSPORTATION.INSURANCE", "Seguro vehículo",  "TRANSPORTATION", "EXPENSE_FIXED",     None),
    CategorySeed("TRANSPORTATION.LICENSING", "Licencias",        "TRANSPORTATION", "EXPENSE_FIXED",     None),
    CategorySeed("TRANSPORTATION.MAINTENANCE","Mantenimiento",   "TRANSPORTATION", "EXPENSE_VARIABLE",  None),

    CategorySeed("ENTERTAINMENT.NETFLIX",    "Netflix",          "ENTERTAINMENT",  "EXPENSE_FIXED",     329.0),
    CategorySeed("ENTERTAINMENT.HBO",        "HBO",              "ENTERTAINMENT",  "EXPENSE_FIXED",      85.0),
    CategorySeed("ENTERTAINMENT.PRIME",      "Amazon Prime",     "ENTERTAINMENT",  "EXPENSE_FIXED",      99.0),
    CategorySeed("ENTERTAINMENT.DISNEY",     "Disney+ y Star",   "ENTERTAINMENT",  "EXPENSE_FIXED",     None),
    CategorySeed("ENTERTAINMENT.SPOTIFY",    "Spotify",          "ENTERTAINMENT",  "EXPENSE_FIXED",     179.0),
    CategorySeed("ENTERTAINMENT.HAWAIANO",   "Hawaiano",         "ENTERTAINMENT",  "EXPENSE_VARIABLE",  None),
    CategorySeed("ENTERTAINMENT.DIVERSION",  "Diversión",        "ENTERTAINMENT",  "EXPENSE_VARIABLE",  None),
    CategorySeed("ENTERTAINMENT.MELI_PLUS",  "Meli+",            "ENTERTAINMENT",  "EXPENSE_FIXED",     130.0),
    CategorySeed("ENTERTAINMENT.YOUTUBE_PREMIUM", "YouTube Premium", "ENTERTAINMENT", "EXPENSE_FIXED",  179.0),

    CategorySeed("FOOD.COMIDA",              "Comida",           "FOOD",           "EXPENSE_VARIABLE", 3000.0),
    CategorySeed("FOOD.DESPENSA",            "Despensa",         "FOOD",           "EXPENSE_VARIABLE", 1500.0),
    CategorySeed("FOOD.LIMPIEZA",            "Limpieza",         "FOOD",           "EXPENSE_VARIABLE",  None),

    CategorySeed("PETS.COMIDA",              "Comida gatas",     "PETS",           "EXPENSE_FIXED",     980.0),
    CategorySeed("PETS.GROOMING",            "Grooming",         "PETS",           "EXPENSE_VARIABLE",  None),
    CategorySeed("PETS.VETERINARIO",         "Veterinario",      "PETS",           "EXPENSE_VARIABLE",  None),

    CategorySeed("LOANS.COPPEL",             "Coppel",           "LOANS",          "EXPENSE_INSTALLMENT", None),
    CategorySeed("LOANS.LIVERPOOL",          "Liverpool",        "LOANS",          "EXPENSE_INSTALLMENT", None),
    CategorySeed("LOANS.SEARS",              "Sears",            "LOANS",          "EXPENSE_INSTALLMENT", None),
    CategorySeed("LOANS.WALMART",            "Walmart",          "LOANS",          "EXPENSE_INSTALLMENT", None),
    CategorySeed("LOANS.BANAMEX_CC",         "Banamex Clásica",  "LOANS",          "EXPENSE_INSTALLMENT", None),
    CategorySeed("LOANS.MERCADO_LIBRE",      "Mercado Libre",    "LOANS",          "EXPENSE_INSTALLMENT", None),
    CategorySeed("LOANS.MERCADO_PAGO",       "Mercado Pago",     "LOANS",          "EXPENSE_INSTALLMENT", None),
    CategorySeed("LOANS.OMAR",               "Préstamo Omar",    "LOANS",          "EXPENSE_INSTALLMENT", None),
    CategorySeed("LOANS.DIDI",               "Didi Card",        "LOANS",          "EXPENSE_INSTALLMENT", None),
    CategorySeed("LOANS.BURO",               "Buró de crédito",  "LOANS",          "EXPENSE_INSTALLMENT", None),
    CategorySeed("LOANS.KLAR",               "Klar",             "LOANS",          "EXPENSE_INSTALLMENT", None),

    CategorySeed("SEGUROS_MEDICOS.BENJI",    "Seguro Benji",     "SEGUROS_MEDICOS","EXPENSE_FIXED",     None),
    CategorySeed("SEGUROS_MEDICOS.NORMA",    "Seguro Norma",     "SEGUROS_MEDICOS","EXPENSE_FIXED",     None),
    CategorySeed("SEGUROS_MEDICOS.HIJOS",    "Seguro hijos",     "SEGUROS_MEDICOS","EXPENSE_FIXED",     None),
    CategorySeed("SEGUROS_MEDICOS.SANTI",    "Seguro Santi",     "SEGUROS_MEDICOS","EXPENSE_FIXED",     None),

    CategorySeed("TRANSFERENCIAS_FAMILIARES.DAVID",    "Mesada David",    "TRANSFERENCIAS_FAMILIARES", "TRANSFER_INTRA_HOUSEHOLD", None),
    CategorySeed("TRANSFERENCIAS_FAMILIARES.PAU",      "Mesada Normita",  "TRANSFERENCIAS_FAMILIARES", "TRANSFER_INTRA_HOUSEHOLD", None),
    CategorySeed("TRANSFERENCIAS_FAMILIARES.SANTIAGO", "Mesada Santiago", "TRANSFERENCIAS_FAMILIARES", "TRANSFER_INTRA_HOUSEHOLD", None),
    CategorySeed("TRANSFERENCIAS_FAMILIARES.NORMA",    "Mesada Norma",    "TRANSFERENCIAS_FAMILIARES", "TRANSFER_INTRA_HOUSEHOLD", None),
    CategorySeed("TRANSFERENCIAS_FAMILIARES.AGUSTIN",  "Mesada Agustín",  "TRANSFERENCIAS_FAMILIARES", "TRANSFER_INTRA_HOUSEHOLD", None),
    CategorySeed("TRANSFERENCIAS_FAMILIARES.COCHE",    "Transferencia coche", "TRANSFERENCIAS_FAMILIARES", "TRANSFER_INTRA_HOUSEHOLD", None),
    CategorySeed("TRANSFERENCIAS_FAMILIARES.INSCRIPCIONES", "Inscripciones", "TRANSFERENCIAS_FAMILIARES", "TRANSFER_INTRA_HOUSEHOLD", None),

    CategorySeed("ESCUELA.DAVID",            "Colegiatura David",    "ESCUELA",    "EXPENSE_FIXED",     None),
    CategorySeed("ESCUELA.PAU",              "Colegiatura Normita",  "ESCUELA",    "EXPENSE_FIXED",     None),
    CategorySeed("ESCUELA.SANTIAGO",          "Colegiatura Santiago", "ESCUELA",    "EXPENSE_FIXED",     None),

    CategorySeed("SAVINGS.EMPRESA",          "Ahorro Empresa",   "SAVINGS",        "SAVINGS",           None),
    CategorySeed("SAVINGS.TARJETA",          "Tarjeta de ahorro","SAVINGS",        "SAVINGS",           None),
    CategorySeed("SAVINGS.RETIREMENT",       "Retiro",           "SAVINGS",        "SAVINGS",           None),
    CategorySeed("SAVINGS.INVESTMENT",       "Inversión",        "SAVINGS",        "SAVINGS",           None),
    CategorySeed("SAVINGS.EFECTIVO",         "Ahorro efectivo",  "SAVINGS",        "SAVINGS",           None),
    # Caja Chica es un ahorro para todos (confirmado por Agustín 2026-07-07),
    # no un gasto "Otros" — vive bajo SAVINGS.
    CategorySeed("SAVINGS.CAJA_CHICA",       "Caja Chica",       "SAVINGS",        "SAVINGS",           None),

    CategorySeed("SERVICIOS_EXTERNOS.ARACELI","Araceli",         "SERVICIOS_EXTERNOS", "EXPENSE_FIXED",  None),
    CategorySeed("SERVICIOS_EXTERNOS.PSICOLOGA","Psicóloga",     "SERVICIOS_EXTERNOS", "EXPENSE_FIXED",  None),
    CategorySeed("SERVICIOS_EXTERNOS.BERNARDO","Bernardo",       "SERVICIOS_EXTERNOS", "EXPENSE_VARIABLE", None),
    CategorySeed("SERVICIOS_EXTERNOS.CONTADOR","Contador",       "SERVICIOS_EXTERNOS", "EXPENSE_VARIABLE", None),

    CategorySeed("OTHER.TELEFONO",           "Teléfono celular", "OTHER",          "EXPENSE_FIXED",     None),
    CategorySeed("OTHER.SALUD",              "Salud (psicóloga)","OTHER",          "EXPENSE_FIXED",     900.0),
    CategorySeed("OTHER.HAWAIANO",           "Clases hawaiano", "OTHER",          "EXPENSE_VARIABLE",  None),
    CategorySeed("OTHER.ADOBE",              "Adobe Creative",  "OTHER",          "EXPENSE_FIXED",     249.0),
    CategorySeed("OTHER.GOOGLE_ONE",         "Google One",       "OTHER",          "EXPENSE_FIXED",      39.0),
    CategorySeed("OTHER.GOOGLE_HOME",        "Google Home",      "OTHER",          "EXPENSE_FIXED",     200.0),
    CategorySeed("OTHER.KIGO",               "Kigo",             "OTHER",          "EXPENSE_FIXED",     100.0),
    CategorySeed("OTHER.COURSERA",           "Coursera",         "OTHER",          "EXPENSE_FIXED",     288.0),
    CategorySeed("OTHER.MARY",               "Mary",             "OTHER",          "EXPENSE_VARIABLE",  None),

    CategorySeed("INGRESOS.SUELDO",          "Sueldo",           "INGRESOS",       "INCOME",            None),
)

CATEGORIES_BY_CODE: dict[str, CategorySeed] = {c.code: c for c in CATEGORIES}


# ── 1.4 Mapeo encabezado Excel → código de categoría raíz ─────────────────────
#
# Las hojas contienen estos 14 rótulos en MAYÚSCULAS. Se normalizan quitando
# tildes/espacios y se asocian a la categoría raíz correspondiente.
EXCEL_SECTION_TO_CATEGORY: dict[str, str] = {
    "HOUSING":                "HOUSING",
    "TRANSPORTATION":         "TRANSPORTATION",
    "PERSONAL":               "SEGUROS_MEDICOS",          # rebautizado en spec
    "PERSONAL CARE":          "PERSONAL_CARE",
    "FOOD":                   "FOOD",
    "PETS":                   "PETS",
    "ENTERTAINMENT":          "ENTERTAINMENT",
    "LOANS":                  "LOANS",
    "SCHOOL":                 "TRANSFERENCIAS_FAMILIARES", # rebautizado
    "TAXES":                  "TRANSFERENCIAS_FAMILIARES", # rebautizado
    "SAVINGS OR INVESTMENTS": "SAVINGS",
    "GIFTS AND DONATIONS":    "GIFTS",
    "LEGAL":                  "LEGAL",
    "OTHERS":                 "OTHER",
}


# ══════════════════════════════════════════════════════════════════════════════
#  BLOQUE 2 · ESQUEMA DE BASE DE DATOS (DDL compatible con Room)
# ══════════════════════════════════════════════════════════════════════════════
#
#  Las sentencias CREATE TABLE replican exactamente lo que Room genera a partir
#  de las anotaciones @Entity, @ColumnInfo, @ForeignKey y @Index.
#
#  Convenciones de mapeo Kotlin → SQL:
#     String        → TEXT
#     Int, Boolean  → INTEGER          (Boolean: 0/1)
#     Long          → INTEGER
#     Double        → REAL
#     Nullable      → se omite NOT NULL
#
#  Los índices se crean aparte porque Room los genera con el prefijo
#  "index_<tabla>_<col1>_<col2>".
# ══════════════════════════════════════════════════════════════════════════════

SCHEMA_DDL: tuple[str, ...] = (
    # ── household ────────────────────────────────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS `household` (
        `id` TEXT NOT NULL,
        `name` TEXT NOT NULL,
        `currency` TEXT NOT NULL,
        `timezone` TEXT NOT NULL,
        `quincena_anchor` TEXT NOT NULL,
        `created_at` INTEGER NOT NULL,
        PRIMARY KEY(`id`)
    )
    """,
    "CREATE UNIQUE INDEX IF NOT EXISTS `index_household_name` ON `household` (`name`)",

    # ── member ───────────────────────────────────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS `member` (
        `id` TEXT NOT NULL,
        `household_id` TEXT NOT NULL,
        `display_name` TEXT NOT NULL,
        `short_aliases` TEXT NOT NULL,
        `role` TEXT NOT NULL,
        `is_active` INTEGER NOT NULL,
        `default_income_mxn` REAL,
        `meta` TEXT NOT NULL,
        PRIMARY KEY(`id`),
        FOREIGN KEY(`household_id`) REFERENCES `household`(`id`) ON UPDATE NO ACTION ON DELETE CASCADE
    )
    """,
    "CREATE INDEX IF NOT EXISTS `index_member_household_id_is_active` ON `member` (`household_id`, `is_active`)",
    "CREATE UNIQUE INDEX IF NOT EXISTS `index_member_household_id_display_name` ON `member` (`household_id`, `display_name`)",

    # ── category ─────────────────────────────────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS `category` (
        `id` TEXT NOT NULL,
        `household_id` TEXT NOT NULL,
        `parent_id` TEXT,
        `code` TEXT NOT NULL,
        `display_name` TEXT NOT NULL,
        `icon` TEXT,
        `color_hex` TEXT,
        `kind` TEXT NOT NULL,
        `budget_default_mxn` REAL,
        `sort_order` INTEGER NOT NULL,
        PRIMARY KEY(`id`),
        FOREIGN KEY(`household_id`) REFERENCES `household`(`id`) ON UPDATE NO ACTION ON DELETE NO ACTION,
        FOREIGN KEY(`parent_id`)    REFERENCES `category`(`id`)  ON UPDATE NO ACTION ON DELETE SET NULL
    )
    """,
    "CREATE UNIQUE INDEX IF NOT EXISTS `index_category_household_id_code` ON `category` (`household_id`, `code`)",
    "CREATE INDEX IF NOT EXISTS `index_category_parent_id` ON `category` (`parent_id`)",

    # ── payment_method ───────────────────────────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS `payment_method` (
        `id` TEXT NOT NULL,
        `household_id` TEXT NOT NULL,
        `display_name` TEXT NOT NULL,
        `kind` TEXT NOT NULL,
        `issuer` TEXT,
        `last4` TEXT,
        `cutoff_day` INTEGER,
        `due_day` INTEGER,
        `credit_limit_mxn` REAL,
        `current_balance_mxn` REAL NOT NULL,
        `interest_apr` REAL,
        `owner_member_id` TEXT,
        `is_active` INTEGER NOT NULL,
        PRIMARY KEY(`id`),
        FOREIGN KEY(`household_id`)    REFERENCES `household`(`id`) ON UPDATE NO ACTION ON DELETE NO ACTION,
        FOREIGN KEY(`owner_member_id`) REFERENCES `member`(`id`)    ON UPDATE NO ACTION ON DELETE SET NULL
    )
    """,
    "CREATE INDEX IF NOT EXISTS `index_payment_method_household_id` ON `payment_method` (`household_id`)",
    "CREATE INDEX IF NOT EXISTS `index_payment_method_owner_member_id` ON `payment_method` (`owner_member_id`)",

    # ── quincena ─────────────────────────────────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS `quincena` (
        `id` TEXT NOT NULL,
        `household_id` TEXT NOT NULL,
        `year` INTEGER NOT NULL,
        `month` INTEGER NOT NULL,
        `half` TEXT NOT NULL,
        `start_date` TEXT NOT NULL,
        `end_date` TEXT NOT NULL,
        `label` TEXT NOT NULL,
        `projected_income_mxn` REAL NOT NULL,
        `projected_expenses_mxn` REAL NOT NULL,
        `actual_income_mxn` REAL NOT NULL,
        `actual_expenses_mxn` REAL NOT NULL,
        `status` TEXT NOT NULL,
        `closed_at` INTEGER,
        PRIMARY KEY(`id`),
        FOREIGN KEY(`household_id`) REFERENCES `household`(`id`) ON UPDATE NO ACTION ON DELETE NO ACTION
    )
    """,
    "CREATE UNIQUE INDEX IF NOT EXISTS `index_quincena_household_id_year_month_half` ON `quincena` (`household_id`, `year`, `month`, `half`)",
    "CREATE INDEX IF NOT EXISTS `index_quincena_household_id_start_date_end_date` ON `quincena` (`household_id`, `start_date`, `end_date`)",
    "CREATE INDEX IF NOT EXISTS `index_quincena_household_id_status` ON `quincena` (`household_id`, `status`)",

    # ── income_source ────────────────────────────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS `income_source` (
        `id` TEXT NOT NULL,
        `household_id` TEXT NOT NULL,
        `quincena_id` TEXT NOT NULL,
        `member_id` TEXT NOT NULL,
        `label` TEXT NOT NULL,
        `amount_mxn` REAL NOT NULL,
        `cadence` TEXT NOT NULL,
        `expected_date` TEXT NOT NULL,
        `payment_method_id` TEXT,
        `status` TEXT NOT NULL,
        `created_at` INTEGER NOT NULL,
        PRIMARY KEY(`id`),
        FOREIGN KEY(`household_id`)      REFERENCES `household`(`id`)       ON UPDATE NO ACTION ON DELETE NO ACTION,
        FOREIGN KEY(`quincena_id`)       REFERENCES `quincena`(`id`)        ON UPDATE NO ACTION ON DELETE NO ACTION,
        FOREIGN KEY(`member_id`)         REFERENCES `member`(`id`)          ON UPDATE NO ACTION ON DELETE NO ACTION,
        FOREIGN KEY(`payment_method_id`) REFERENCES `payment_method`(`id`)  ON UPDATE NO ACTION ON DELETE SET NULL
    )
    """,
    "CREATE INDEX IF NOT EXISTS `index_income_source_household_id` ON `income_source` (`household_id`)",
    "CREATE INDEX IF NOT EXISTS `index_income_source_quincena_id` ON `income_source` (`quincena_id`)",
    "CREATE INDEX IF NOT EXISTS `index_income_source_member_id` ON `income_source` (`member_id`)",
    "CREATE INDEX IF NOT EXISTS `index_income_source_payment_method_id` ON `income_source` (`payment_method_id`)",

    # ── recurrence_template ──────────────────────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS `recurrence_template` (
        `id` TEXT NOT NULL,
        `household_id` TEXT NOT NULL,
        `concept` TEXT NOT NULL,
        `category_id` TEXT NOT NULL,
        `default_amount_mxn` REAL NOT NULL,
        `default_payment_method_id` TEXT,
        `cadence` TEXT NOT NULL,
        `cadence_detail` TEXT NOT NULL,
        `next_expected_date` TEXT,
        `default_beneficiary_ids` TEXT NOT NULL,
        `default_payer_split` TEXT NOT NULL,
        `is_active` INTEGER NOT NULL,
        `confidence_score` REAL NOT NULL,
        `learned_from_expense_ids` TEXT NOT NULL,
        PRIMARY KEY(`id`),
        FOREIGN KEY(`household_id`)              REFERENCES `household`(`id`)       ON UPDATE NO ACTION ON DELETE NO ACTION,
        FOREIGN KEY(`category_id`)               REFERENCES `category`(`id`)        ON UPDATE NO ACTION ON DELETE NO ACTION,
        FOREIGN KEY(`default_payment_method_id`) REFERENCES `payment_method`(`id`)  ON UPDATE NO ACTION ON DELETE SET NULL
    )
    """,
    "CREATE INDEX IF NOT EXISTS `index_recurrence_template_household_id_is_active` ON `recurrence_template` (`household_id`, `is_active`)",
    "CREATE INDEX IF NOT EXISTS `index_recurrence_template_category_id` ON `recurrence_template` (`category_id`)",
    "CREATE INDEX IF NOT EXISTS `index_recurrence_template_default_payment_method_id` ON `recurrence_template` (`default_payment_method_id`)",

    # ── installment_plan ─────────────────────────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS `installment_plan` (
        `id` TEXT NOT NULL,
        `household_id` TEXT NOT NULL,
        `display_name` TEXT NOT NULL,
        `creditor_member_id` TEXT,
        `payment_method_id` TEXT,
        `principal_mxn` REAL NOT NULL,
        `total_installments` INTEGER NOT NULL,
        `installment_amount_mxn` REAL NOT NULL,
        `interest_rate_apr` REAL,
        `start_date` TEXT NOT NULL,
        `current_installment` INTEGER NOT NULL,
        `status` TEXT NOT NULL,
        `category_id` TEXT,
        PRIMARY KEY(`id`),
        FOREIGN KEY(`household_id`)       REFERENCES `household`(`id`)       ON UPDATE NO ACTION ON DELETE NO ACTION,
        FOREIGN KEY(`creditor_member_id`) REFERENCES `member`(`id`)          ON UPDATE NO ACTION ON DELETE SET NULL,
        FOREIGN KEY(`payment_method_id`)  REFERENCES `payment_method`(`id`)  ON UPDATE NO ACTION ON DELETE SET NULL,
        FOREIGN KEY(`category_id`)        REFERENCES `category`(`id`)        ON UPDATE NO ACTION ON DELETE SET NULL
    )
    """,
    "CREATE INDEX IF NOT EXISTS `index_installment_plan_household_id_status` ON `installment_plan` (`household_id`, `status`)",
    "CREATE INDEX IF NOT EXISTS `index_installment_plan_creditor_member_id` ON `installment_plan` (`creditor_member_id`)",
    "CREATE INDEX IF NOT EXISTS `index_installment_plan_payment_method_id` ON `installment_plan` (`payment_method_id`)",
    "CREATE INDEX IF NOT EXISTS `index_installment_plan_category_id` ON `installment_plan` (`category_id`)",

    # ── expense ──────────────────────────────────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS `expense` (
        `id` TEXT NOT NULL,
        `household_id` TEXT NOT NULL,
        `occurred_at` INTEGER NOT NULL,
        `quincena_id` TEXT NOT NULL,
        `category_id` TEXT NOT NULL,
        `concept` TEXT NOT NULL,
        `amount_mxn` REAL NOT NULL,
        `payment_method_id` TEXT NOT NULL,
        `recurrence_template_id` TEXT,
        `installment_plan_id` TEXT,
        `installment_number` INTEGER,
        `installment_principal_mxn` REAL,
        `installment_interest_mxn` REAL,
        `status` TEXT NOT NULL,
        `notes` TEXT,
        `created_at` INTEGER NOT NULL,
        `created_by_member_id` TEXT,
        PRIMARY KEY(`id`),
        FOREIGN KEY(`household_id`)           REFERENCES `household`(`id`)            ON UPDATE NO ACTION ON DELETE NO ACTION,
        FOREIGN KEY(`quincena_id`)            REFERENCES `quincena`(`id`)             ON UPDATE NO ACTION ON DELETE NO ACTION,
        FOREIGN KEY(`category_id`)            REFERENCES `category`(`id`)             ON UPDATE NO ACTION ON DELETE NO ACTION,
        FOREIGN KEY(`payment_method_id`)      REFERENCES `payment_method`(`id`)       ON UPDATE NO ACTION ON DELETE NO ACTION,
        FOREIGN KEY(`recurrence_template_id`) REFERENCES `recurrence_template`(`id`)  ON UPDATE NO ACTION ON DELETE SET NULL,
        FOREIGN KEY(`installment_plan_id`)    REFERENCES `installment_plan`(`id`)     ON UPDATE NO ACTION ON DELETE SET NULL,
        FOREIGN KEY(`created_by_member_id`)   REFERENCES `member`(`id`)               ON UPDATE NO ACTION ON DELETE SET NULL
    )
    """,
    "CREATE INDEX IF NOT EXISTS `index_expense_quincena_id_status` ON `expense` (`quincena_id`, `status`)",
    "CREATE INDEX IF NOT EXISTS `index_expense_household_id_occurred_at` ON `expense` (`household_id`, `occurred_at`)",
    "CREATE INDEX IF NOT EXISTS `index_expense_category_id_occurred_at` ON `expense` (`category_id`, `occurred_at`)",
    "CREATE INDEX IF NOT EXISTS `index_expense_payment_method_id` ON `expense` (`payment_method_id`)",
    "CREATE INDEX IF NOT EXISTS `index_expense_recurrence_template_id` ON `expense` (`recurrence_template_id`)",
    "CREATE INDEX IF NOT EXISTS `index_expense_installment_plan_id` ON `expense` (`installment_plan_id`)",
    "CREATE INDEX IF NOT EXISTS `index_expense_created_by_member_id` ON `expense` (`created_by_member_id`)",

    # ── expense_attribution ──────────────────────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS `expense_attribution` (
        `id` TEXT NOT NULL,
        `expense_id` TEXT NOT NULL,
        `member_id` TEXT NOT NULL,
        `role` TEXT NOT NULL,
        `share_bps` INTEGER NOT NULL,
        `share_amount_mxn` REAL NOT NULL,
        PRIMARY KEY(`id`),
        FOREIGN KEY(`expense_id`) REFERENCES `expense`(`id`) ON UPDATE NO ACTION ON DELETE CASCADE,
        FOREIGN KEY(`member_id`)  REFERENCES `member`(`id`)  ON UPDATE NO ACTION ON DELETE NO ACTION
    )
    """,
    "CREATE INDEX IF NOT EXISTS `index_expense_attribution_expense_id_role` ON `expense_attribution` (`expense_id`, `role`)",
    "CREATE INDEX IF NOT EXISTS `index_expense_attribution_member_id_role` ON `expense_attribution` (`member_id`, `role`)",

    # ── loan ─────────────────────────────────────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS `loan` (
        `id` TEXT NOT NULL,
        `household_id` TEXT NOT NULL,
        `debtor_member_id` TEXT NOT NULL,
        `principal_mxn` REAL NOT NULL,
        `remaining_balance_mxn` REAL NOT NULL,
        `agreed_interest_mxn` REAL NOT NULL,
        `issued_at` TEXT NOT NULL,
        `due_at` TEXT,
        `payment_schedule_id` TEXT,
        `notes` TEXT,
        PRIMARY KEY(`id`),
        FOREIGN KEY(`household_id`)        REFERENCES `household`(`id`)         ON UPDATE NO ACTION ON DELETE NO ACTION,
        FOREIGN KEY(`debtor_member_id`)    REFERENCES `member`(`id`)            ON UPDATE NO ACTION ON DELETE NO ACTION,
        FOREIGN KEY(`payment_schedule_id`) REFERENCES `installment_plan`(`id`)  ON UPDATE NO ACTION ON DELETE SET NULL
    )
    """,
    "CREATE INDEX IF NOT EXISTS `index_loan_household_id` ON `loan` (`household_id`)",
    "CREATE INDEX IF NOT EXISTS `index_loan_debtor_member_id` ON `loan` (`debtor_member_id`)",
    "CREATE INDEX IF NOT EXISTS `index_loan_payment_schedule_id` ON `loan` (`payment_schedule_id`)",

    # ── savings_goal ─────────────────────────────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS `savings_goal` (
        `id` TEXT NOT NULL,
        `household_id` TEXT NOT NULL,
        `name` TEXT NOT NULL,
        `target_mxn` REAL NOT NULL,
        `current_mxn` REAL NOT NULL,
        `target_date` TEXT,
        `linked_payment_method_id` TEXT,
        PRIMARY KEY(`id`),
        FOREIGN KEY(`household_id`)             REFERENCES `household`(`id`)       ON UPDATE NO ACTION ON DELETE NO ACTION,
        FOREIGN KEY(`linked_payment_method_id`) REFERENCES `payment_method`(`id`)  ON UPDATE NO ACTION ON DELETE SET NULL
    )
    """,
    "CREATE INDEX IF NOT EXISTS `index_savings_goal_household_id` ON `savings_goal` (`household_id`)",
    "CREATE INDEX IF NOT EXISTS `index_savings_goal_linked_payment_method_id` ON `savings_goal` (`linked_payment_method_id`)",

    # ── metadatos Android / Room ─────────────────────────────────────────────
    "CREATE TABLE IF NOT EXISTS `android_metadata` (`locale` TEXT DEFAULT 'es_MX')",
    """
    CREATE TABLE IF NOT EXISTS `room_master_table` (
        `id` INTEGER PRIMARY KEY,
        `identity_hash` TEXT
    )
    """,
)


# ══════════════════════════════════════════════════════════════════════════════
#  BLOQUE 3 · UTILIDADES DE LIMPIEZA DE DATOS
# ══════════════════════════════════════════════════════════════════════════════

def _normalize(text: str) -> str:
    """Mayúsculas + sin tildes + sin espacios múltiples. Usado para matching."""
    if text is None:
        return ""
    # Unicode NFD descompone letras con tilde en letra + combining mark; luego
    # se descartan los diacríticos.
    nfkd = unicodedata.normalize("NFKD", str(text))
    stripped = "".join(ch for ch in nfkd if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", stripped).strip().upper()


def _as_float(value) -> Optional[float]:
    """
    Convierte una celda arbitraria a ``float`` en MXN. Tolera:
      * ``None``, ``""`` → ``None``
      * ``0``, ``0.0`` → ``None``  (ruido estructural del Excel)
      * strings tipo "1,234.50", "$3 200", " 1500 " → float limpio
      * fórmulas rotas (#REF!, #DIV/0!, #VALUE!) → ``None``
      * booleanos → ``None`` (no son montos)
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        f = float(value)
        return f if f != 0 else None
    s = str(value).strip()
    if not s or s.startswith("#"):
        return None
    # Quitar símbolos de moneda y separadores de miles.
    cleaned = re.sub(r"[^\d.\-]", "", s.replace(",", ""))
    try:
        f = float(cleaned)
        return f if f != 0 else None
    except ValueError:
        return None


def _as_str(value) -> Optional[str]:
    """Colapsa a string limpio o ``None`` si la celda es ruido."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


# ══════════════════════════════════════════════════════════════════════════════
#  BLOQUE 4 · PARSER DE NOMBRES DE HOJA → (año, mes, quincena)
# ══════════════════════════════════════════════════════════════════════════════

# Mapeo robusto de nombres de mes (con y sin acento, abreviado o completo) → 1-12
_MONTH_TOKENS: dict[str, int] = {
    "ENE": 1, "ENERO": 1,
    "FEB": 2, "FEBRERO": 2, "FEBRREO": 2, "FEBRERERO": 2,  # typos reales del Excel
    "MAR": 3, "MARZO": 3,
    "ABR": 4, "ABRIL": 4,
    "MAY": 5, "MAYO": 5,
    "JUN": 6, "JUNIO": 6,
    "JUL": 7, "JULIO": 7,
    "AGO": 8, "AGOSTO": 8,
    "SEP": 9, "SEPTIEMBRE": 9, "SEPT": 9,
    "OCT": 10, "OCTUBRE": 10,
    "NOV": 11, "NOVIEMBRE": 11,
    "DIC": 12, "DICIEMBRE": 12,
}

# Regex tolerante al caos. Captura: (día_inicio, día_fin, mes_token, año opcional)
# Acepta: "1 al 15 Abril", "1al 15 Mayo", "16 al 31 de Enero 2026",
#         "Quincena 16 al 28 de febrereo", "16 al 30 junio)"
_SHEET_NAME_RE = re.compile(
    r"""
    (?:Quincena|Quin\.?|Quin)?         # prefijo opcional
    \s*
    (?P<start>\d{1,2})                  # día inicio
    \s*(?:al|a)\s*
    (?P<end>\d{1,2})                    # día fin
    \s*(?:de\s*)?
    (?P<month>[A-Za-zÁÉÍÓÚáéíóúñÑ]+)    # nombre del mes (tolera typos, tildes)
    \s*
    (?P<year>\d{4})?                    # año opcional
    """,
    re.VERBOSE | re.IGNORECASE,
)


@dataclass(frozen=True)
class SheetPeriod:
    """Período normalizado extraído de un nombre de hoja."""
    year: int
    month: int
    half: str              # "FIRST" (1-15) | "SECOND" (16-fin)
    start_date: date
    end_date: date
    source_name: str       # nombre original de la hoja, para trazabilidad

    @property
    def label(self) -> str:
        meses = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        q = "Q1" if self.half == "FIRST" else "Q2"
        return f"{q} {meses[self.month]} {self.year}"

    @property
    def quincena_id(self) -> str:
        return did(f"quincena:{self.year:04d}-{self.month:02d}-{self.half}")


def parse_sheet_name(name: str, fallback_year: int) -> Optional[SheetPeriod]:
    """
    Convierte un nombre de hoja a un ``SheetPeriod`` normalizado.

    * El año, si no aparece en el nombre, se toma de ``fallback_year``.
    * Si el día final del nombre excede los días reales del mes (p. ej.
      "31 septiembre"), se recorta al último día real del mes.
    """
    match = _SHEET_NAME_RE.search(name)
    if not match:
        LOG.warning("No se pudo parsear el nombre de hoja: %r", name)
        return None

    start = int(match.group("start"))
    end = int(match.group("end"))
    month_token = _normalize(match.group("month"))
    year = int(match.group("year")) if match.group("year") else fallback_year

    month = _MONTH_TOKENS.get(month_token)
    if month is None:
        # Último recurso: probar los primeros 3 caracteres normalizados.
        month = _MONTH_TOKENS.get(month_token[:3])
    if month is None:
        LOG.warning("Mes irreconocible %r en hoja %r", month_token, name)
        return None

    # Número real de días del mes, corrigiendo finales imposibles (p. ej. 31 sep).
    last_day = calendar.monthrange(year, month)[1]
    if end > last_day:
        LOG.debug("Recortando día fin %d → %d en hoja %r", end, last_day, name)
        end = last_day

    half = "FIRST" if start <= 15 else "SECOND"

    # Las quincenas SECOND cubren hasta el último día REAL del mes aunque el
    # nombre de la hoja diga "al 30" en un mes de 31 (mar/may/jul/ago/dic):
    # el autor del Excel usa "16 al 30" como plantilla fija. Sin esta
    # extensión, el día 31 quedaría fuera de toda quincena y los gastos
    # distribuidos por `occurred_at` nunca caerían en él.
    if half == "SECOND" and end < last_day:
        LOG.debug("Extendiendo día fin %d → %d (fin de mes real) en hoja %r",
                  end, last_day, name)
        end = last_day
    return SheetPeriod(
        year=year,
        month=month,
        half=half,
        start_date=date(year, month, start),
        end_date=date(year, month, end),
        source_name=name,
    )


def build_sheet_periods(sheet_names: list[str]) -> dict[str, SheetPeriod]:
    """
    Devuelve un dict ``sheet_name → SheetPeriod`` iterando con un cursor de año
    que empieza en 2025 y avanza a 2026 en cuanto una hoja lo especifica.
    """
    periods: dict[str, SheetPeriod] = {}
    running_year = 2025
    last_month_seen = 0

    for name in sheet_names:
        # Si el nombre de la hoja contiene "2026" lo usamos directo; si no, se
        # hereda el running_year, con detección de wrap-around mes→mes descendente.
        explicit_year = re.search(r"(20\d{2})", name)
        if explicit_year:
            running_year = int(explicit_year.group(1))

        period = parse_sheet_name(name, fallback_year=running_year)
        if not period:
            continue

        # Heurística de wrap-around para hojas 2025 sin año explícito:
        # si el mes retrocede con respecto al anterior, se asume salto de año.
        if (not explicit_year
                and period.month < last_month_seen
                and last_month_seen - period.month > 6):
            running_year += 1
            period = parse_sheet_name(name, fallback_year=running_year) or period

        periods[name] = period
        last_month_seen = period.month

    return periods


# ══════════════════════════════════════════════════════════════════════════════
#  BLOQUE 5 · ESCÁNER DE BLOQUES DE CATEGORÍA EN UNA HOJA
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class BudgetLine:
    """Una fila presupuestaria extraída de una hoja."""
    sheet_name: str
    section_code: str             # código de categoría raíz (HOUSING, FOOD, …)
    concept: str                  # descripción libre tal cual aparece en Excel
    projected: float              # Projected Cost > 0
    paid_by_norma: Optional[float]     # columna D o I
    paid_by_benjamin: Optional[float]  # columna E o J


@dataclass
class SectionHeader:
    """Posición de un header de categoría dentro de la hoja."""
    row: int
    name_col: int                 # columna donde está el nombre (B o G)
    projected_col: int            # columna "Projected Cost" (C o H)
    norma_col: int                # columna "Norma" / "Actual Cost" (D o I)
    benjamin_col: int             # columna "Benjamin" / "Difference" (E o J)
    section_code: str             # código canónico mapeado


def find_section_headers(ws: Worksheet) -> list[SectionHeader]:
    """
    Escanea las primeras ~65 filas en busca de celdas cuyo valor contenga
    "Projected Cost"; por cada una, el header de la categoría está una
    columna a la izquierda y la fila de datos empieza en la siguiente fila.
    """
    headers: list[SectionHeader] = []
    max_row = min(ws.max_row, 65)
    max_col = min(ws.max_column, 14)

    for r in range(1, max_row + 1):
        for c in range(2, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if not isinstance(val, str):
                continue
            if "Projected Cost" not in val:
                continue
            name_cell = ws.cell(row=r, column=c - 1).value
            if not name_cell:
                continue
            section_norm = _normalize(str(name_cell))
            section_code = EXCEL_SECTION_TO_CATEGORY.get(section_norm)
            if section_code is None:
                # No es una sección conocida; lo registramos pero no lo usamos
                # para evitar importar ruido (p. ej. "Column1" o celdas sueltas).
                LOG.debug("Sección desconocida ignorada: %r (hoja %r)",
                          name_cell, ws.title)
                continue
            headers.append(SectionHeader(
                row=r,
                name_col=c - 1,
                projected_col=c,
                norma_col=c + 1,
                benjamin_col=c + 2,
                section_code=section_code,
            ))
    return headers


# Tokens de terminación de bloque: al encontrarlos, la sección se da por cerrada.
_BLOCK_TERMINATORS = {"SUBTOTAL", "TOTAL", "TOTAL SUBTOTAL"}


def extract_budget_lines(ws: Worksheet) -> list[BudgetLine]:
    """
    Recorre todas las secciones de una hoja y emite una ``BudgetLine`` por
    cada renglón con ``projected > 0``. El escaneo se detiene al tocar una
    fila "Subtotal" o una celda vacía prolongada.
    """
    lines: list[BudgetLine] = []
    for hdr in find_section_headers(ws):
        # Descender desde la fila siguiente hasta encontrar terminador o un
        # gap de 3 filas vacías consecutivas (protección contra loops).
        empty_streak = 0
        for r in range(hdr.row + 1, min(hdr.row + 25, ws.max_row + 1)):
            concept_cell = ws.cell(row=r, column=hdr.name_col).value
            projected = _as_float(ws.cell(row=r, column=hdr.projected_col).value)
            norma = _as_float(ws.cell(row=r, column=hdr.norma_col).value)
            benja = _as_float(ws.cell(row=r, column=hdr.benjamin_col).value)

            concept_str = _as_str(concept_cell) or ""
            concept_upper = _normalize(concept_str)

            if concept_upper in _BLOCK_TERMINATORS:
                break
            if not concept_str and projected is None:
                empty_streak += 1
                if empty_streak >= 3:
                    break
                continue
            empty_streak = 0

            # Solo nos interesan filas con un monto presupuestado real.
            if projected is None or projected <= 0:
                continue
            if not concept_str or concept_upper in {"OTHER"}:
                # "Other" en filas intermedias es ruido típico del template.
                continue

            lines.append(BudgetLine(
                sheet_name=ws.title,
                section_code=hdr.section_code,
                concept=concept_str,
                projected=projected,
                paid_by_norma=norma,
                paid_by_benjamin=benja,
            ))
    return lines


# ══════════════════════════════════════════════════════════════════════════════
#  BLOQUE 6 · MOTOR DE REGLAS + RESOLVERS SEMÁNTICOS
# ══════════════════════════════════════════════════════════════════════════════

# ── 6.0 Cargador de reglas de atribución ─────────────────────────────────────

@dataclass(frozen=True)
class AttributionRule:
    """Una regla de override cargada desde attribution_rules.json."""
    rule_id: str
    concept_pattern: str           # Patrón normalizado (UPPER, sin tildes)
    match_type: str                # 'contains' | 'exact_normalized'
    section_filter: Optional[str]  # None = aplica a cualquier sección
    override_category: Optional[str]
    override_beneficiaries: Optional[list[str]]  # member keys
    override_payment_method: Optional[str]       # payment_method key or None


def load_attribution_rules(rules_path: Path) -> list[AttributionRule]:
    """Lee attribution_rules.json y devuelve las reglas en orden de prioridad."""
    if not rules_path.exists():
        LOG.warning("No se encontro archivo de reglas: %s — usando heuristicas puras.", rules_path)
        return []
    with open(rules_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    rules = []
    for r in data.get("rules", []):
        rules.append(AttributionRule(
            rule_id=r.get("id", "unknown"),
            concept_pattern=r["concept_pattern"],
            match_type=r.get("match_type", "contains"),
            section_filter=r.get("section_filter"),
            override_category=r.get("override_category"),
            override_beneficiaries=r.get("override_beneficiaries"),
            override_payment_method=r.get("override_payment_method"),
        ))
    LOG.info("Cargadas %d reglas de atribucion desde %s", len(rules), rules_path)
    return rules


def match_rule(line: BudgetLine, rules: list[AttributionRule]) -> Optional[AttributionRule]:
    """
    Evalua las reglas en orden y devuelve la primera que matchea.
    Retorna None si ninguna aplica (se usa el fallback heuristico).
    """
    concept_norm = _normalize(line.concept)
    for rule in rules:
        # Filtro de seccion
        if rule.section_filter is not None and rule.section_filter != line.section_code:
            continue
        # Match de concepto
        if rule.match_type == "exact_normalized":
            if concept_norm != rule.concept_pattern:
                continue
        elif rule.match_type == "contains":
            if rule.concept_pattern not in concept_norm:
                continue
        else:
            continue  # tipo desconocido
        return rule
    return None


# Variable global que se inicializa en EtlPipeline.run()
_ATTRIBUTION_RULES: list[AttributionRule] = []


# ── 6.1 Resolvers con soporte de reglas ──────────────────────────────────────

def resolve_category_code(line: BudgetLine) -> str:
    """
    Devuelve el ``code`` de la categoria resuelta. Prioridad:
      1. Regla explicita de attribution_rules.json
      2. Matching heuristico contra hojas hijas
      3. Fallback: raiz de la seccion

    Se expone el CODE (no el id) porque la capa de canonicalizacion de
    conceptos condiciona sus mapeos por la categoria RESUELTA (el mismo
    token "DAVID"/"PAU"/"SANTIAGO" cae en categorias distintas segun la
    seccion del Excel).
    """
    # 1. Regla explicita
    rule = match_rule(line, _ATTRIBUTION_RULES)
    if rule and rule.override_category:
        code = rule.override_category
        if code in CATEGORIES_BY_CODE:
            return code
        # Si la regla referencia una categoria que no existe, log y fallback
        LOG.warning("Regla '%s' referencia categoria inexistente '%s'", rule.rule_id, code)

    # 2. Matching heuristico contra hojas hijas
    concept_norm = _normalize(line.concept)
    for cat in CATEGORIES:
        if cat.parent_code != line.section_code:
            continue
        leaf = cat.code.rsplit(".", 1)[-1]
        leaf_display = _normalize(cat.display_name)
        if leaf in concept_norm or leaf_display in concept_norm:
            return cat.code

    # 3. Fallback: raiz de la seccion.
    return line.section_code


def resolve_category(line: BudgetLine) -> str:
    """Devuelve el ``id`` de la categoria resuelta (wrapper de conveniencia)."""
    return CATEGORIES_BY_CODE[resolve_category_code(line)].id


def resolve_payment_method(line: BudgetLine) -> str:
    """
    Resuelve el metodo de pago. Prioridad:
      1. Regla explicita de attribution_rules.json
      2. Heuristica basada en concepto y columnas Norma/Benjamin
    """
    # 1. Regla explicita
    rule = match_rule(line, _ATTRIBUTION_RULES)
    if rule and rule.override_payment_method:
        pm_key = rule.override_payment_method
        if pm_key in PAYMENT_METHODS_BY_KEY:
            return PAYMENT_METHODS_BY_KEY[pm_key].id

    # 2. Heuristica basada en concepto
    concept_norm = _normalize(line.concept)
    wallet_keywords: list[tuple[str, str]] = [
        ("BANAMEX CLASICA", "banamex_cc"),
        ("BANAMEX DEBITO",  "banamex_deb"),
        ("MERCADO LIBRE",   "mercado_libre"),
        ("MERCADO PAGO",    "mercado_pago"),
        ("COPPEL",          "coppel"),
        ("LIVERPOOL",       "liverpool"),
        ("SEARS",           "sears"),
        ("WALMART",         "walmart"),
        ("KLAR",            "klar"),
        ("BBVA",            "bbva"),
        ("AHORRO EMPRESA",  "ahorro_empresa"),
    ]
    for needle, key in wallet_keywords:
        if needle in concept_norm:
            return PAYMENT_METHODS_BY_KEY[key].id

    # Quien paga determina el default — la mayoria paga con efectivo
    return PAYMENT_METHODS_BY_KEY["efectivo"].id


def resolve_beneficiaries(line: BudgetLine) -> list[str]:
    """
    Devuelve la lista de ``member.id`` beneficiarios. Prioridad:
      1. Regla explicita de attribution_rules.json
      2. Heuristica por alias en el concepto
      3. Fallback: adultos
    """
    # 1. Regla explicita
    rule = match_rule(line, _ATTRIBUTION_RULES)
    if rule and rule.override_beneficiaries:
        ids = []
        for key in rule.override_beneficiaries:
            if key in MEMBERS_BY_KEY:
                ids.append(MEMBERS_BY_KEY[key].id)
            else:
                LOG.warning("Regla '%s' referencia miembro inexistente '%s'", rule.rule_id, key)
        if ids:
            return ids

    # 2. Heuristica por alias
    concept_norm = _normalize(line.concept)
    hits: list[str] = []
    for m in MEMBERS:
        if m.role not in ("BENEFICIARY_DEPENDENT", "PAYER_ADULT"):
            continue
        for alias in m.aliases:
            if _normalize(alias) in concept_norm:
                hits.append(m.id)
                break

    # Deduplica preservando orden.
    seen = set()
    uniq = [h for h in hits if not (h in seen or seen.add(h))]

    if uniq:
        return uniq

    # 3. Fallback: adultos como beneficiarios.
    return [MEMBERS_BY_KEY["norma"].id, MEMBERS_BY_KEY["benjamin"].id]


def split_basis_points(count: int) -> list[int]:
    """
    Divide 10_000 basis points en ``count`` porciones tan iguales como sea
    posible. El residuo se suma a la última porción para garantizar la
    invariante ``sum == 10_000``.
    """
    if count <= 0:
        raise ValueError("count debe ser > 0")
    base = 10_000 // count
    shares = [base] * count
    shares[-1] += 10_000 - (base * count)
    return shares


# ── 6.2 Canonicalizacion de conceptos ────────────────────────────────────────
#
# El Excel nombra el MISMO cargo de formas distintas entre hojas ("Didi",
# "Didi Card", "Didi tarjeta"; "Cochecito"/"Gasolina chochecito"; "David
# Abril"/"Inscripcion David"; "Prestamo Omar 3..10"). Esta capa renombra el
# concepto a una forma canonica ANTES de insertar, preservando el texto crudo
# del Excel en `notes` ("Concepto original Excel: <crudo>") para trazabilidad.
#
# IMPORTANTE: los mapeos que dependen de nombres de hijos se condicionan por
# la categoria RESUELTA (o la seccion), no solo por el texto: el token
# "DAVID" es colegiatura en TRANSFERENCIAS, telefono en HOUSING, etc.
# Las reglas se evaluan en orden: la primera que matchea gana.

PAYER_RULE_CUTOVER = date(2025, 10, 1)  # ver _derive_payer_shares


@dataclass(frozen=True)
class CanonicalConceptRule:
    """Renombra un concepto crudo del Excel a su forma canonica."""
    pattern: str                          # regex FULLMATCH sobre concepto normalizado
    canonical: str                        # concepto canonico a persistir
    category_filter: Optional[str] = None # code de la categoria RESUELTA
    section_filter: Optional[str] = None  # seccion raiz del Excel
    note: Optional[str] = None            # nota extra; admite {n} de grupos del regex


CANONICAL_CONCEPT_RULES: tuple[CanonicalConceptRule, ...] = (
    # Personas / servicios externos.
    CanonicalConceptRule(r"BERNA(RDO)?", "Bernardo"),

    # DiDi: la tarjeta de credito DiDi aparece como "Didi", "Didi Card" y
    # "Didi tarjeta" — es el mismo cargo.
    CanonicalConceptRule(r"DIDI( CARD| TARJETA)?", "DiDi"),

    # Escuela: colegiaturas/inscripciones/libros con mes o beneficiario
    # embebido en el concepto. Se condiciona por categoria resuelta para no
    # tocar los "DAVID"/"PAU"/"SANTIAGO" de otras secciones.
    CanonicalConceptRule(r".*", "Escuela David",    category_filter="ESCUELA.DAVID"),
    CanonicalConceptRule(r".*", "Escuela Santiago", category_filter="ESCUELA.SANTIAGO"),
    CanonicalConceptRule(r".*", "Escuela Normita",  category_filter="ESCUELA.PAU"),

    # Transferencias a Normita (Pau) que NO son colegiatura (el "Pau" de $800
    # en OTHER y "Pau Raul" — Raul es el novio de Normita, el gasto es de ella).
    CanonicalConceptRule(r"PAU RAUL", "Normita",
                         note="gasto de Normita (Pau); Raúl es su novio"),
    CanonicalConceptRule(r"PAU", "Normita",
                         category_filter="TRANSFERENCIAS_FAMILIARES.PAU"),

    # Seguro de los tres hijos mayores (sección PERSONAL del Excel).
    CanonicalConceptRule(r"PAU,? DAVID,? AGUS", "Normita, David y Agus"),

    # Typos y tildes.
    CanonicalConceptRule(r"HAWAI?[NI]?ANO", "Hawaiano"),   # HAWAIANO / HAWAINANO
    CanonicalConceptRule(r"DIVERSION", "Diversión"),

    # Cochecito: "COCHECITO" (quincenas 1-15) y "GASOLINA CHOCHECITO"
    # (quincenas 16-fin) son el MISMO cargo de gasolina con nombre
    # inconsistente — verificado DISJUNTOS por quincena en el Excel
    # (2026-07-07: nunca coinciden en la misma hoja). Se unifican.
    CanonicalConceptRule(r"(GASOLINA CHOCHECITO|COCHECITO)", "Gasolina Cochecito",
                         section_filter="TRANSPORTATION"),
    # El COCHECITO de $8,000 en OTHERS es un servicio/reparacion, no gasolina.
    CanonicalConceptRule(r"COCHECITO", "Cochecito (servicio)",
                         section_filter="OTHER"),

    # Prestamo Omar: serie de pagos numerados en el concepto.
    CanonicalConceptRule(r"PRESTAMO OMAR (?P<n>\d+)", "Préstamo Omar",
                         note="pago {n} de la serie"),
    CanonicalConceptRule(r"PRESTAMO (?P<n>\d+) \(?OMAR\)?", "Préstamo Omar",
                         note="pago {n} de la serie"),

    # MSI / series con "N de M" embebido. NO se crea installment_plan
    # (montos inconsistentes entre pagos); queda anotado como mejora futura.
    CanonicalConceptRule(r"MERCADO LIBRE (?P<n>\d+) DE 12", "Mercado Libre MSI",
                         note="pago {n} de 12"),
    CanonicalConceptRule(r"BURO DE CREDITO (?P<n>\d+) DE 3", "Buró de Crédito",
                         note="pago {n} de 3"),

    # Telefonos por persona (tildes + el "David" de $249 suelto en HOUSING,
    # que la regla housing_david_solo ya manda a OTHER.TELEFONO).
    CanonicalConceptRule(r"DAVID", "Teléfono David", category_filter="OTHER.TELEFONO"),
    CanonicalConceptRule(r"TELEFONO DAVID", "Teléfono David"),
    CanonicalConceptRule(r"TELEFONO NORMA", "Teléfono Norma"),
    CanonicalConceptRule(r"TELEFONO SANTI", "Teléfono Santi"),
    CanonicalConceptRule(r"TELEFONO BENJI", "Teléfono Benji"),
    CanonicalConceptRule(r"TELEFONO PAU", "Teléfono Normita"),

    # Pulido cosmético: capitalización/tildes consistentes para conceptos que
    # el Excel trae en minúsculas/MAYÚSCULAS crudas (una sola variante cada
    # uno; no cambia clasificación, solo presentación).
    CanonicalConceptRule(r"CAJA CHICA", "Caja Chica"),
    CanonicalConceptRule(r"KIGO", "Kigo"),
    CanonicalConceptRule(r"GOOGLE NEST", "Google Home"),
    CanonicalConceptRule(r"MERCADO PAGO", "Mercado Pago"),
    CanonicalConceptRule(r"SUSCRIPCION NIVEL 6", "Suscripción Nivel 6"),
    CanonicalConceptRule(r"MARCO Y OMAR", "Marco y Omar"),
    CanonicalConceptRule(r"PSICOLOGA", "Psicóloga"),
    CanonicalConceptRule(r"BANAMEX CLASICA", "Banamex Clásica"),
    CanonicalConceptRule(r"YOUTUBE", "YouTube"),
)


def canonicalize_concept(
    concept_raw: str,
    section_code: str,
    category_code: str,
) -> tuple[str, Optional[str]]:
    """
    Devuelve ``(concepto_canonico, nota_extra)``. Si ninguna regla matchea,
    el concepto queda tal cual vino del Excel y la nota es ``None``.
    """
    concept_norm = _normalize(concept_raw)
    for rule in CANONICAL_CONCEPT_RULES:
        if rule.section_filter is not None and rule.section_filter != section_code:
            continue
        if rule.category_filter is not None and rule.category_filter != category_code:
            continue
        m = re.fullmatch(rule.pattern, concept_norm)
        if not m:
            continue
        note = None
        if rule.note:
            groups = {k: v for k, v in m.groupdict().items() if v}
            note = rule.note.format(**groups) if groups else rule.note
        return rule.canonical, note
    return concept_raw.strip(), None


# ── 6.3 Split de lineas compartidas (una linea del Excel → N gastos) ─────────
#
# Objetivo (Agustín, 2026-07-07): "por persona un mismo cargo con mismo
# concepto cada mes". Las lineas de telefono que agrupan a varias personas
# ("Telefono Pau y David", "Telefono Movistar") se parten en N gastos hijos,
# uno por persona, con montos repartidos equitativamente (residuo de
# centavos al ultimo — el total exacto se preserva) y beneficiario 100% el
# miembro respectivo. Los ids son deterministas (id del padre + member_key).

@dataclass(frozen=True)
class SplitRule:
    """Parte una linea del Excel en N gastos hijos, uno por miembro."""
    pattern: str                              # regex FULLMATCH sobre concepto normalizado
    section_filter: Optional[str]
    children: tuple[tuple[str, str], ...]     # (member_key, concepto canonico hijo)


SPLIT_RULES: tuple[SplitRule, ...] = (
    SplitRule(r"TELEFONO PAU Y DAVID", "HOUSING", (
        ("pau",   "Teléfono Normita"),
        ("david", "Teléfono David"),
    )),
    SplitRule(r"TELEFONO MOVISTAR", "HOUSING", (
        ("norma",    "Teléfono Norma"),
        ("pau",      "Teléfono Normita"),
        ("david",    "Teléfono David"),
        ("santiago", "Teléfono Santi"),
    )),
)


def match_split_rule(line: BudgetLine) -> Optional[SplitRule]:
    concept_norm = _normalize(line.concept)
    for rule in SPLIT_RULES:
        if rule.section_filter is not None and rule.section_filter != line.section_code:
            continue
        if re.fullmatch(rule.pattern, concept_norm):
            return rule
    return None


def split_amount_mxn(total: float, count: int) -> list[float]:
    """
    Reparte ``total`` en ``count`` montos iguales trabajando en centavos;
    el residuo va al ultimo para que la suma sea EXACTAMENTE el total.
    """
    if count <= 0:
        raise ValueError("count debe ser > 0")
    total_cents = round(total * 100)
    base = total_cents // count
    cents = [base] * count
    cents[-1] += total_cents - base * count
    return [c / 100.0 for c in cents]


# ══════════════════════════════════════════════════════════════════════════════
#  BLOQUE 6.4 · PLANTILLAS DE RECURRENCIA + PLANES MSI (modelo de futuro)
# ══════════════════════════════════════════════════════════════════════════════
#
# Agustín (2026-07-07): la semilla ya no siembra el futuro como gastos PLANNED.
# En su lugar, las OBLIGACIONES FIJAS recurrentes se modelan como
# `recurrence_template` (la app materializa PLANNED en runtime cada quincena) y
# las series a meses reales como `installment_plan`.
#
# Conceptos a templatizar: una plantilla por cada obligación fija recurrente.
# Se EXCLUYEN los gastos variables (comida, despensa, limpieza, gasolinas,
# diversión). Los conceptos se identifican por su forma CANÓNICA ya persistida
# en `expense.concept` (ver CANONICAL_CONCEPT_RULES / SPLIT_RULES).
#
# NOTA sobre "Normita, David y Agus": la lista de Agustín menciona "Normita" y
# "David y Agus (seguro)" por separado, pero en el Excel/semilla es UN solo
# concepto de seguro de los tres hijos mayores ("Normita, David y Agus"), así
# que produce UNA plantilla (36 en total, no 37).
RECURRENCE_TEMPLATE_CONCEPTS: tuple[str, ...] = (
    "Hipoteca", "Internet", "Agua", "Electricidad",
    "Netflix", "Prime", "HBO", "Spotify", "YouTube",
    "Adobe", "Google", "Google Home", "Kigo", "Coursera",
    "Suscripción Nivel 6",
    "Teléfono Norma", "Teléfono Normita", "Teléfono David",
    "Teléfono Santi", "Teléfono Benji",
    "Walmart", "Banamex Clásica", "Liverpool", "Sears", "Coppel",
    "Mercado Pago", "DiDi", "Klar",
    "Comida Gatas", "Psicóloga", "Préstamo Omar",
    "Benji", "Normita, David y Agus",
    "Escuela David", "Escuela Santiago", "Escuela Normita",
)


# ══════════════════════════════════════════════════════════════════════════════
#  BLOQUE 7 · ORQUESTADOR ETL
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class EtlStats:
    """Contadores para el reporte final."""
    sheets_parsed: int = 0
    quincenas_written: int = 0
    expenses_written: int = 0
    attributions_written: int = 0
    incomes_written: int = 0
    skipped_sheets: list[str] = field(default_factory=list)


class EtlPipeline:
    """
    Ejecuta el pipeline completo Excel → SQLite.

    Uso:
        >>> EtlPipeline(excel_path, db_path).run()
    """

    def __init__(self, excel_path: Path, db_path: Path, rules_path: Optional[Path] = None):
        self.excel_path = excel_path
        self.db_path = db_path
        self.rules_path = rules_path
        self.stats = EtlStats()
        # Suma de la celda E9 ("Ahorro empresa") de las hojas ya iniciadas;
        # se vuelca a savings_goal.current_mxn en _finalize.
        self._ahorro_empresa_mxn = 0.0

    # ── punto de entrada ─────────────────────────────────────────────────────

    def run(self) -> EtlStats:
        global _ATTRIBUTION_RULES

        # Cargar reglas de atribucion si se proporcionaron
        if self.rules_path:
            _ATTRIBUTION_RULES = load_attribution_rules(self.rules_path)
        else:
            # Intentar cargar desde el directorio del script
            default_rules = Path(__file__).parent / "attribution_rules.json"
            _ATTRIBUTION_RULES = load_attribution_rules(default_rules)

        LOG.info("Abriendo Excel: %s", self.excel_path)
        wb = load_workbook(self.excel_path, data_only=True, read_only=False)

        LOG.info("Inicializando base de datos: %s", self.db_path)
        if self.db_path.exists():
            self.db_path.unlink()  # reinicio limpio

        with sqlite3.connect(self.db_path) as conn:
            # PRAGMA foreign_keys se activa por sesión — la app lo hace al
            # abrir la DB; aquí lo dejamos desactivado durante la inserción
            # para evitar bloqueos de orden. Se reactivará al final.
            conn.execute("PRAGMA foreign_keys = OFF")
            conn.execute("PRAGMA journal_mode = WAL")

            self._create_schema(conn)
            self._insert_seed(conn)
            self._ingest_sheets(conn, wb)
            self._insert_oneoff_seeds(conn)
            self._seed_recurrence_templates(conn)
            self._seed_installment_plans(conn)
            self._finalize(conn)

        LOG.info("ETL completado: %s", self.stats)
        return self.stats

    # ── paso 1: esquema ──────────────────────────────────────────────────────

    def _create_schema(self, conn: sqlite3.Connection) -> None:
        cur = conn.cursor()
        for stmt in SCHEMA_DDL:
            cur.execute(stmt)
        cur.execute(
            "INSERT OR REPLACE INTO room_master_table (id, identity_hash) VALUES (42, ?)",
            (ROOM_IDENTITY_HASH,),
        )
        cur.execute("INSERT OR REPLACE INTO android_metadata (locale) VALUES ('es_MX')")
        conn.commit()

    # ── paso 2: datos semilla (household, member, category, wallet) ─────────

    def _insert_seed(self, conn: sqlite3.Connection) -> None:
        import json
        cur = conn.cursor()

        cur.execute(
            """INSERT INTO household
                   (id, name, currency, timezone, quincena_anchor, created_at)
               VALUES (?, ?, 'MXN', ?, 'CALENDAR', ?)""",
            (HOUSEHOLD_ID, HOUSEHOLD_NAME, HOUSEHOLD_TZ, NOW_EPOCH_MS),
        )

        for m in MEMBERS:
            cur.execute(
                """INSERT INTO member
                       (id, household_id, display_name, short_aliases, role,
                        is_active, default_income_mxn, meta)
                   VALUES (?, ?, ?, ?, ?, 1, ?, '{}')""",
                (m.id, HOUSEHOLD_ID, m.display_name,
                 json.dumps(list(m.aliases), ensure_ascii=False),
                 m.role, m.default_income),
            )

        # Categorías insertadas en dos pasadas: primero raíces (parent_code None),
        # luego hijas, para respetar la FK parent_id.
        for c in CATEGORIES:
            if c.parent_code is None:
                cur.execute(
                    """INSERT INTO category
                           (id, household_id, parent_id, code, display_name,
                            icon, color_hex, kind, budget_default_mxn, sort_order)
                       VALUES (?, ?, NULL, ?, ?, NULL, NULL, ?, ?, ?)""",
                    (c.id, HOUSEHOLD_ID, c.code, c.display_name,
                     c.kind, c.budget_default, c.sort_order),
                )
        for c in CATEGORIES:
            if c.parent_code is not None:
                parent_id = CATEGORIES_BY_CODE[c.parent_code].id
                cur.execute(
                    """INSERT INTO category
                           (id, household_id, parent_id, code, display_name,
                            icon, color_hex, kind, budget_default_mxn, sort_order)
                       VALUES (?, ?, ?, ?, ?, NULL, NULL, ?, ?, ?)""",
                    (c.id, HOUSEHOLD_ID, parent_id, c.code, c.display_name,
                     c.kind, c.budget_default, c.sort_order),
                )

        for pm in PAYMENT_METHODS:
            owner_id = MEMBERS_BY_KEY[pm.owner_key].id if pm.owner_key else None
            cur.execute(
                """INSERT INTO payment_method
                       (id, household_id, display_name, kind, issuer, last4,
                        cutoff_day, due_day, credit_limit_mxn,
                        current_balance_mxn, interest_apr, owner_member_id,
                        is_active)
                   VALUES (?, ?, ?, ?, ?, NULL, NULL, NULL, NULL, 0.0, NULL, ?, 1)""",
                (pm.id, HOUSEHOLD_ID, pm.display_name, pm.kind, pm.issuer, owner_id),
            )

        # Préstamo histórico a Jaudiel (filas 116-118 del Excel).
        jaudiel_id = MEMBERS_BY_KEY["jaudiel"].id
        cur.execute(
            """INSERT INTO loan
                   (id, household_id, debtor_member_id, principal_mxn,
                    remaining_balance_mxn, agreed_interest_mxn, issued_at,
                    due_at, payment_schedule_id, notes)
               VALUES (?, ?, ?, ?, ?, 0.0, '2025-03-01', NULL, NULL, ?)""",
            (did("loan:jaudiel"), HOUSEHOLD_ID, jaudiel_id,
             105_000.0, 105_000.0, "Préstamo histórico registrado en Excel (R116)"),
        )

        # Meta de ahorro seed vinculada a Ahorro Empresa.
        cur.execute(
            """INSERT INTO savings_goal
                   (id, household_id, name, target_mxn, current_mxn,
                    target_date, linked_payment_method_id)
               VALUES (?, ?, 'Ahorro Empresa anual', 60000.0, 0.0, NULL, ?)""",
            (did("savings_goal:ahorro_empresa"),
             HOUSEHOLD_ID,
             PAYMENT_METHODS_BY_KEY["ahorro_empresa"].id),
        )

        conn.commit()
        LOG.info("Semillas: %d miembros, %d categorías, %d wallets",
                 len(MEMBERS), len(CATEGORIES), len(PAYMENT_METHODS))

    # ── paso 3: ingesta de hojas quincenales ────────────────────────────────

    def _ingest_sheets(self, conn: sqlite3.Connection, wb) -> None:
        sheet_names = wb.sheetnames
        periods = build_sheet_periods(sheet_names)

        # MODELO DE FUTURO (Agustín, 2026-07-07): la semilla NO emite quincenas
        # futuras. Solo se conserva hasta la quincena ACTIVA (la que cubre HOY);
        # las posteriores las crea el rollover de la app al llegar la fecha.
        # `active_end` = último día de la quincena activa: 15 si HOY cae en la
        # primera mitad, o el fin de mes real si cae en la segunda.
        if TODAY.day <= 15:
            active_end = date(TODAY.year, TODAY.month, 15)
        else:
            last_day = calendar.monthrange(TODAY.year, TODAY.month)[1]
            active_end = date(TODAY.year, TODAY.month, last_day)

        for name in sheet_names:
            period = periods.get(name)
            if period is None:
                self.stats.skipped_sheets.append(name)
                continue

            # Descartar hojas cuyo período empieza DESPUÉS de la quincena activa
            # (p. ej. jul-2ª, ago-1ª, ago-2ª cuando HOY = 2026-07-07). No se
            # siembran quincena/ingresos/gastos para el futuro.
            if period.start_date > active_end:
                self.stats.skipped_sheets.append(name)
                continue

            ws = wb[name]
            self._process_sheet(conn, ws, period)
            self.stats.sheets_parsed += 1

    def _process_sheet(
        self,
        conn: sqlite3.Connection,
        ws: Worksheet,
        period: SheetPeriod,
    ) -> None:
        cur = conn.cursor()

        # 3.1 · Extraer montos fijos de ingresos (celdas E4 y E5).
        # Fieles al Excel: si la celda está vacía NO se inventa un fallback al
        # default del miembro — un E4 vacío significa que ese ingreso ya no
        # existe (Benjamín deja de percibir sueldo desde oct-2025) y un E5
        # cambiante refleja los aumentos reales de Norma (60k → 75k → 85k).
        # `_insert_income` ya descarta montos <= 0, así que la quincena queda
        # sin income_source para ese miembro, que es exactamente lo correcto.
        benja_income = _as_float(ws["E4"].value) or 0.0
        norma_income = _as_float(ws["E5"].value) or 0.0
        projected_income = benja_income + norma_income

        # 3.2 · Extraer las líneas de presupuesto (gastos).
        lines = extract_budget_lines(ws)

        projected_expenses = sum(l.projected for l in lines)

        # 3.2b · Precomputar la fecha de cada gasto ANTES de insertar la
        # quincena: la quincena ACTIVE necesita saber cuántos de sus gastos
        # quedarán POSTED (occurred <= hoy) para escribir actual_expenses.
        # occurred_at se distribuye uniformemente dentro del período para que
        # las consultas temporales de la app tengan sentido.
        span_days = max((period.end_date - period.start_date).days, 1)
        dated_lines: list[tuple[BudgetLine, date, int]] = []
        for idx, line in enumerate(lines):
            day_offset = (idx * span_days) // max(len(lines), 1)
            expense_date = date.fromordinal(period.start_date.toordinal() + day_offset)
            occurred_at = int(datetime.combine(
                expense_date,
                time(hour=12),
                tzinfo=timezone.utc,
            ).timestamp() * 1000)
            dated_lines.append((line, expense_date, occurred_at))

        # 3.3 · Insertar Quincena, con status derivado de la fecha de
        # generación (TODAY) en vez del 'CLOSED' histórico:
        #   * terminó antes de hoy      → CLOSED, actual = projected (histórico)
        #   * hoy cae dentro            → ACTIVE, actuales = solo lo ya POSTED
        #   * empieza después de hoy    → PROVISIONED, actuales en cero
        if period.end_date < TODAY:
            status = "CLOSED"
            closed_at: Optional[int] = NOW_EPOCH_MS
            actual_income = projected_income
            actual_expenses = projected_expenses
        elif period.start_date > TODAY:
            status = "PROVISIONED"
            closed_at = None
            actual_income = 0.0
            actual_expenses = 0.0
        else:
            status = "ACTIVE"
            closed_at = None
            # Solo cuenta la mitad ya transcurrida: los gastos cuya fecha
            # distribuida quedará POSTED y los ingresos ya cobrados (la fecha
            # esperada es el inicio del período, que para la ACTIVE ya pasó).
            actual_expenses = sum(
                l.projected for l, expense_date, _ in dated_lines
                if expense_date <= TODAY
            )
            actual_income = sum(
                amount for amount in (benja_income, norma_income)
                if amount > 0 and period.start_date <= TODAY
            )

        cur.execute(
            """INSERT INTO quincena
                   (id, household_id, year, month, half, start_date, end_date,
                    label, projected_income_mxn, projected_expenses_mxn,
                    actual_income_mxn, actual_expenses_mxn, status, closed_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                period.quincena_id,
                HOUSEHOLD_ID,
                period.year,
                period.month,
                period.half,
                period.start_date.isoformat(),
                period.end_date.isoformat(),
                period.label,
                projected_income,
                projected_expenses,
                actual_income,
                actual_expenses,
                status,
                closed_at,
            ),
        )
        self.stats.quincenas_written += 1

        # 3.4 · Insertar fuentes de ingreso (Benjamín + Norma).
        self._insert_income(
            cur, period,
            member_key="benjamin",
            amount=benja_income,
            label="Sueldo Benjamín",
            payment_method_key="ahorro_empresa",  # empresa deposita aquí
        )
        self._insert_income(
            cur, period,
            member_key="norma",
            amount=norma_income,
            label="Sueldo Norma",
            payment_method_key="bbva",
        )

        # 3.5 · Insertar gastos + atribuciones (fechas precomputadas en 3.2b).
        for line, _, occurred_at in dated_lines:
            self._insert_expense(cur, period, line, occurred_at)

        # 3.6 · Acumular el ahorro empresa ya apartado (celda E9): alimenta
        # el current_mxn de la meta "Ahorro Empresa anual" en _finalize.
        # Solo cuentan períodos ya iniciados — lo futuro aún no está apartado.
        if period.start_date <= TODAY:
            self._ahorro_empresa_mxn += _as_float(ws["E9"].value) or 0.0

        conn.commit()

    # ── paso 3a: insertar un IncomeSource ──────────────────────────────────

    def _insert_income(
        self,
        cur: sqlite3.Cursor,
        period: SheetPeriod,
        member_key: str,
        amount: float,
        label: str,
        payment_method_key: str,
    ) -> None:
        if amount <= 0:
            return
        income_id = did(f"income:{period.quincena_id}:{member_key}")
        # POSTED solo si el ingreso ya debió cobrarse (expected_date = inicio
        # del período <= hoy); los de quincenas futuras nacen PLANNED.
        status = "POSTED" if period.start_date <= TODAY else "PLANNED"
        cur.execute(
            """INSERT INTO income_source
                   (id, household_id, quincena_id, member_id, label,
                    amount_mxn, cadence, expected_date, payment_method_id,
                    status, created_at)
               VALUES (?, ?, ?, ?, ?, ?, 'QUINCENAL', ?, ?, ?, ?)""",
            (
                income_id,
                HOUSEHOLD_ID,
                period.quincena_id,
                MEMBERS_BY_KEY[member_key].id,
                label,
                amount,
                period.start_date.isoformat(),
                PAYMENT_METHODS_BY_KEY[payment_method_key].id,
                status,
                NOW_EPOCH_MS,
            ),
        )
        self.stats.incomes_written += 1

    # ── paso 3b: insertar un Expense + sus atribuciones ────────────────────

    def _insert_expense(
        self,
        cur: sqlite3.Cursor,
        period: SheetPeriod,
        line: BudgetLine,
        occurred_at: int,
    ) -> None:
        # POSTED solo si el gasto ya "ocurrió" respecto a la fecha de
        # generación; lo futuro (quincenas PROVISIONED y la mitad restante de
        # la ACTIVE) nace PLANNED, que es como la app modela pagos pendientes
        # confirmables. La fecha se lee del propio occurred_at en UTC porque
        # así se generó (mediodía UTC del día distribuido).
        occurred_date = datetime.fromtimestamp(
            occurred_at / 1000, tz=timezone.utc
        ).date()
        status = "POSTED" if occurred_date <= TODAY else "PLANNED"

        # CERO gastos PLANNED sembrados (Agustín, 2026-07-07): la semilla solo
        # persiste lo ya ocurrido (POSTED). Los pagos futuros de la quincena
        # activa (p. ej. jul 8-15) NO se materializan aquí — los genera la app
        # en runtime desde las plantillas de recurrencia (recurrence_template).
        # Esto también cubre los hijos de split, porque el split se resuelve
        # después de este punto.
        if status != "POSTED":
            return

        category_code = resolve_category_code(line)
        category_id = CATEGORIES_BY_CODE[category_code].id
        wallet_id = resolve_payment_method(line)

        # ¿La linea se parte en N gastos hijos (uno por persona)?
        split = match_split_rule(line)
        if split is not None:
            self._insert_split_children(
                cur, period, line, occurred_at, occurred_date, status,
                category_id, wallet_id, split,
            )
            return

        expense_id = did(
            f"expense:{period.quincena_id}:{line.section_code}:"
            f"{_normalize(line.concept)}"
        )

        # Canonicalizacion: renombra el concepto y preserva el crudo en notes.
        concept, extra_note = canonicalize_concept(
            line.concept, line.section_code, category_code,
        )
        notes = f"Importado desde hoja '{line.sheet_name}'"
        if concept != line.concept.strip():
            notes += f" · Concepto original Excel: {line.concept.strip()}"
        if extra_note:
            notes += f" · {extra_note}"

        cur.execute(
            """INSERT INTO expense
                   (id, household_id, occurred_at, quincena_id, category_id,
                    concept, amount_mxn, payment_method_id,
                    recurrence_template_id, installment_plan_id,
                    installment_number, installment_principal_mxn,
                    installment_interest_mxn, status, notes, created_at,
                    created_by_member_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, NULL, NULL, NULL, NULL, NULL,
                       ?, ?, ?, NULL)""",
            (
                expense_id,
                HOUSEHOLD_ID,
                occurred_at,
                period.quincena_id,
                category_id,
                concept[:64],
                line.projected,
                wallet_id,
                status,
                notes,
                NOW_EPOCH_MS,
            ),
        )
        self.stats.expenses_written += 1

        # Atribuciones BENEFICIARY.
        beneficiaries = resolve_beneficiaries(line)
        bps = split_basis_points(len(beneficiaries))
        for member_id, share in zip(beneficiaries, bps):
            attr_id = did(f"attr:{expense_id}:BEN:{member_id}")
            cur.execute(
                """INSERT INTO expense_attribution
                       (id, expense_id, member_id, role, share_bps, share_amount_mxn)
                   VALUES (?, ?, ?, 'BENEFICIARY', ?, ?)""",
                (attr_id, expense_id, member_id, share,
                 round(line.projected * share / 10_000, 2)),
            )
            self.stats.attributions_written += 1

        # Atribuciones PAYER.
        self._insert_payer_attributions(
            cur, expense_id, line, line.projected, occurred_date, concept,
        )

    def _insert_split_children(
        self,
        cur: sqlite3.Cursor,
        period: SheetPeriod,
        line: BudgetLine,
        occurred_at: int,
        occurred_date: date,
        status: str,
        category_id: str,
        wallet_id: str,
        split: SplitRule,
    ) -> None:
        """
        Emite N gastos hijos a partir de una linea compartida del Excel.
        Cada hijo: id determinista (padre + member_key), monto equitativo
        (residuo de centavos al ultimo — el total exacto se preserva),
        beneficiario 100% el miembro respectivo, y nota de trazabilidad.
        """
        amounts = split_amount_mxn(line.projected, len(split.children))
        for (member_key, child_concept), amount in zip(split.children, amounts):
            member_id = MEMBERS_BY_KEY[member_key].id
            child_id = did(
                f"expense:{period.quincena_id}:{line.section_code}:"
                f"{_normalize(line.concept)}:{member_key}"
            )
            notes = (
                f"Importado desde hoja '{line.sheet_name}' · Parte de línea "
                f"compartida del Excel: {line.concept.strip()} "
                f"(${line.projected:,.2f})"
            )
            cur.execute(
                """INSERT INTO expense
                       (id, household_id, occurred_at, quincena_id, category_id,
                        concept, amount_mxn, payment_method_id,
                        recurrence_template_id, installment_plan_id,
                        installment_number, installment_principal_mxn,
                        installment_interest_mxn, status, notes, created_at,
                        created_by_member_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, NULL, NULL, NULL, NULL, NULL,
                           ?, ?, ?, NULL)""",
                (
                    child_id,
                    HOUSEHOLD_ID,
                    occurred_at,
                    period.quincena_id,
                    category_id,
                    child_concept[:64],
                    amount,
                    wallet_id,
                    status,
                    notes,
                    NOW_EPOCH_MS,
                ),
            )
            self.stats.expenses_written += 1

            # BENEFICIARY: 100% el miembro del hijo.
            attr_id = did(f"attr:{child_id}:BEN:{member_id}")
            cur.execute(
                """INSERT INTO expense_attribution
                       (id, expense_id, member_id, role, share_bps, share_amount_mxn)
                   VALUES (?, ?, ?, 'BENEFICIARY', 10000, ?)""",
                (attr_id, child_id, member_id, amount),
            )
            self.stats.attributions_written += 1

            # PAYER: misma regla que el padre, prorrateada al monto del hijo.
            self._insert_payer_attributions(
                cur, child_id, line, amount, occurred_date, child_concept,
            )

    def _insert_payer_attributions(
        self,
        cur: sqlite3.Cursor,
        expense_id: str,
        line: BudgetLine,
        amount: float,
        expense_date: date,
        concept: str,
    ) -> None:
        payer_shares = self._derive_payer_shares(line, amount, expense_date, concept)
        for member_id, share, share_amount in payer_shares:
            attr_id = did(f"attr:{expense_id}:PAY:{member_id}")
            cur.execute(
                """INSERT INTO expense_attribution
                       (id, expense_id, member_id, role, share_bps, share_amount_mxn)
                   VALUES (?, ?, ?, 'PAYER', ?, ?)""",
                (attr_id, expense_id, member_id, share, share_amount),
            )
            self.stats.attributions_written += 1

    def _derive_payer_shares(
        self,
        line: BudgetLine,
        amount: float,
        expense_date: date,
        concept: str,
    ) -> list[tuple[str, int, float]]:
        """
        Calcula el reparto PAYER. Retorna lista de
        ``(member_id, share_bps, share_amount_mxn)``.

        REGLA DE NEGOCIO (confirmada por Agustín 2026-07-07): Benjamín no
        percibe sueldo desde oct-2025 (la celda E4 del Excel está vacía desde
        entonces), así que todo gasto con fecha >= 2025-10-01 lo paga Norma
        al 100%, con la ÚNICA excepción de Spotify, que Benjamín sigue
        pagando de su bolsillo. Antes de esa fecha el reparto se deriva fiel
        de las columnas Norma/Benjamín del Excel.
        """
        norma_id = MEMBERS_BY_KEY["norma"].id
        benja_id = MEMBERS_BY_KEY["benjamin"].id

        if expense_date >= PAYER_RULE_CUTOVER:
            if "SPOTIFY" in _normalize(concept):
                return [(benja_id, 10_000, amount)]
            return [(norma_id, 10_000, amount)]

        # Los montos del Excel ocasionalmente llegan en negativo cuando el
        # autor anotó un "ajuste" o devolución (p.ej. Benjamin = -200). Para
        # efectos del reparto PAYER esos valores no representan una fracción
        # real del pago — se clampean a 0. Conservar el signo propagaría el
        # error a la división n/total y rompería el invariante de 10_000 bps.
        n = max(line.paid_by_norma or 0.0, 0.0)
        b = max(line.paid_by_benjamin or 0.0, 0.0)
        total = n + b

        if total <= 0:
            # Asume Norma al 100% — es la pagadora principal del hogar.
            return [(norma_id, 10_000, amount)]

        # Normaliza por el total declarado (no por `projected`): las columnas
        # del Excel a veces suman un poco más o un poco menos que el Projected
        # por redondeo manual. Lo que importa es conservar la PROPORCIÓN entre
        # Norma y Benjamín; el contrato de Room exige que share_amount_mxn =
        # amount_mxn * share_bps / 10000, así que se re-deriva aquí y NUNCA
        # se toma el valor bruto del Excel (evita el bug de bps > 10000).
        n_bps = int(round(n / total * 10_000))
        b_bps = 10_000 - n_bps  # garantiza invariante: n_bps + b_bps == 10_000

        result: list[tuple[str, int, float]] = []
        if n_bps > 0:
            result.append((norma_id, n_bps,
                           round(amount * n_bps / 10_000, 2)))
        if b_bps > 0:
            result.append((benja_id, b_bps,
                           round(amount * b_bps / 10_000, 2)))
        return result

    # ── paso 4: cierre ──────────────────────────────────────────────────────

    # ── paso 3.7: seeds one-off fuera de las secciones presupuestales ───────

    def _insert_oneoff_seeds(self, conn: sqlite3.Connection) -> None:
        """
        Registros confirmados por Agustín (2026-07-07) que viven en las NOTAS
        del Excel, no en las secciones presupuestales:

        1. Fondo de $59,000 de Q1 feb-2025 (filas 69-75): aguinaldo de Norma
           gastado fuera del presupuesto. Se registra el ingreso y los dos
           gastos que NO duplican líneas presupuestadas (Mueble $19,900 y
           Reparación coche $6,900); los demás renglones del fondo
           (Liverpool/Sears/Despensa/Pau) ya existen como líneas del
           presupuesto. Los actuales de la quincena se ajustan (dejan de ser
           == projected: hubo gasto extra real cubierto por el aguinaldo).
        2. Cuenta por cobrar "Deben $3,600 — Medicina + rotafolio" (oct-2025),
           reducida a $1,900 en dic-2025/ene-2026 → loan por cobrar con
           deudor "Por identificar" (Norma lo renombra en la app).
        """
        cur = conn.cursor()
        q_feb = did("quincena:2025-02-FIRST")
        norma = MEMBERS_BY_KEY["norma"]
        efectivo_id = PAYMENT_METHODS_BY_KEY["efectivo"].id

        # 1a. Ingreso del aguinaldo (en efectivo: el fondo se siguió en las
        # notas "Ahorro Efectivo", no en un banco).
        cur.execute(
            """INSERT INTO income_source
                   (id, household_id, quincena_id, member_id, label,
                    amount_mxn, cadence, expected_date, payment_method_id,
                    status, created_at)
               VALUES (?, ?, ?, ?, 'Aguinaldo', 59000.0, 'UNICO',
                       '2025-02-01', ?, 'POSTED', ?)""",
            (did(f"income:{q_feb}:aguinaldo"), HOUSEHOLD_ID, q_feb,
             norma.id, efectivo_id, NOW_EPOCH_MS),
        )
        self.stats.incomes_written += 1

        # 1b. Gastos one-off pagados del aguinaldo.
        occurred = int(datetime(2025, 2, 8, 12, tzinfo=timezone.utc)
                       .timestamp() * 1000)
        fondo_nota = ("Pagado del aguinaldo de $59,000 (fondo de feb-2025, "
                      "notas del Excel filas 69-75)")
        oneoffs = (
            ("mueble", "Mueble", 19_900.0, "HOUSING.MUEBLES",
             ["benjamin", "norma", "pau", "david", "agustin", "santiago"]),
            ("reparacion_coche", "Reparación coche", 6_900.0,
             "TRANSPORTATION.MAINTENANCE", ["agustin", "david", "pau"]),
        )
        for key, concepto, monto, cat_code, bene_keys in oneoffs:
            expense_id = did(f"expense:oneoff:2025-02:{key}")
            cur.execute(
                """INSERT INTO expense
                       (id, household_id, occurred_at, quincena_id, category_id,
                        concept, amount_mxn, payment_method_id,
                        recurrence_template_id, installment_plan_id,
                        installment_number, installment_principal_mxn,
                        installment_interest_mxn, status, notes, created_at,
                        created_by_member_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, NULL, NULL, NULL, NULL,
                           NULL, 'POSTED', ?, ?, NULL)""",
                (expense_id, HOUSEHOLD_ID, occurred, q_feb,
                 CATEGORIES_BY_CODE[cat_code].id, concepto, monto,
                 efectivo_id, fondo_nota, NOW_EPOCH_MS),
            )
            self.stats.expenses_written += 1
            bene_ids = [MEMBERS_BY_KEY[k].id for k in bene_keys]
            for member_id, share in zip(bene_ids,
                                        split_basis_points(len(bene_ids))):
                cur.execute(
                    """INSERT INTO expense_attribution
                           (id, expense_id, member_id, role, share_bps,
                            share_amount_mxn)
                       VALUES (?, ?, ?, 'BENEFICIARY', ?, ?)""",
                    (did(f"attr:{expense_id}:BEN:{member_id}"), expense_id,
                     member_id, share, round(monto * share / 10_000, 2)),
                )
                self.stats.attributions_written += 1
            # Pagado por Norma (de su aguinaldo).
            cur.execute(
                """INSERT INTO expense_attribution
                       (id, expense_id, member_id, role, share_bps,
                        share_amount_mxn)
                   VALUES (?, ?, ?, 'PAYER', 10000, ?)""",
                (did(f"attr:{expense_id}:PAY:{norma.id}"), expense_id,
                 norma.id, monto),
            )
            self.stats.attributions_written += 1

        # 1c. Ajustar actuales de la quincena (CLOSED venía con
        # actual == projected; el fondo fue movimiento real extra).
        cur.execute(
            """UPDATE quincena
               SET actual_income_mxn = actual_income_mxn + 59000.0,
                   actual_expenses_mxn = actual_expenses_mxn + 26800.0
               WHERE id = ?""",
            (q_feb,),
        )

        # 2. Cuenta por cobrar "Deben".
        cur.execute(
            """INSERT INTO loan
                   (id, household_id, debtor_member_id, principal_mxn,
                    remaining_balance_mxn, agreed_interest_mxn, issued_at,
                    due_at, payment_schedule_id, notes)
               VALUES (?, ?, ?, 3600.0, 1900.0, 0.0, '2025-10-01', NULL,
                       NULL, ?)""",
            (did("loan:deben_medicina_rotafolio"), HOUSEHOLD_ID,
             MEMBERS_BY_KEY["deudor_pendiente"].id,
             "Cuenta por cobrar de las notas del Excel: 'Deben $3,600 — "
             "Medicina + rotafolio' (oct-2025); para dic-2025/ene-2026 la "
             "nota baja a $1,900. Deudor por identificar/renombrar en la app."),
        )
        conn.commit()

    # ── paso 3.8: plantillas de recurrencia (obligaciones fijas) ────────────

    @staticmethod
    def _next_expected_date(cadence: str, day: int, today: date) -> Optional[str]:
        """
        Próxima fecha futura (> today) según cadencia + día típico. Genera las
        ocurrencias candidatas de los próximos ~5 meses y devuelve la primera
        posterior a hoy en ISO yyyy-MM-dd. Formato de cadencia = enum que parsea
        RecurrenceMaterializer (QUINCENAL_EVERY / MONTHLY_SPECIFIC_HALF).
        """
        cands: list[date] = []
        for i in range(5):
            mm = today.month + i
            yy = today.year + (mm - 1) // 12
            mm = (mm - 1) % 12 + 1
            last = calendar.monthrange(yy, mm)[1]
            if cadence == "QUINCENAL_EVERY":
                cands.append(date(yy, mm, min(max(day, 1), 15)))
                cands.append(date(yy, mm, min(max(day, 16), last)))
            else:  # MONTHLY_SPECIFIC_HALF
                cands.append(date(yy, mm, min(max(day, 1), last)))
        future = [d for d in cands if d > today]
        return min(future).isoformat() if future else None

    def _seed_recurrence_templates(self, conn: sqlite3.Connection) -> None:
        """
        Puebla `recurrence_template` con UNA plantilla por obligación fija
        recurrente (RECURRENCE_TEMPLATE_CONCEPTS). Deriva monto (mediana),
        categoría/wallet/beneficiarios/pagador (del gasto más reciente) y
        cadencia (del patrón de mitades). Solo lee gastos POSTED ya sembrados.

        Contrato JSON (verificado contra RecurrenceMaterializer.kt):
          * cadence: "QUINCENAL_EVERY" (≈2/mes, ambas mitades) |
                     "MONTHLY_SPECIFIC_HALF" (≈1/mes, la mitad del día).
          * cadence_detail: {"day_of_month": <1-31>}. El materializador lo
            coacciona a 1-15 (primera mitad) / 16-fin (segunda) según cadencia.
          * default_beneficiary_ids: JSON array ["<member_id>", ...] (reparto
            equitativo — formato histórico que parsea parseBeneficiaries()).
          * default_payer_split: JSON object {"<member_id>": <bps>} (parsePayerSplit()).
        """
        cur = conn.cursor()
        created = 0
        for concept in RECURRENCE_TEMPLATE_CONCEPTS:
            rows = cur.execute(
                """SELECT id, occurred_at, amount_mxn, category_id, payment_method_id
                   FROM expense
                   WHERE concept = ? AND status = 'POSTED'
                   ORDER BY occurred_at""",
                (concept,),
            ).fetchall()
            if not rows:
                LOG.warning("Plantilla omitida: sin gastos POSTED para %r", concept)
                continue

            amounts = sorted(r[2] for r in rows)
            median_amount = amounts[len(amounts) // 2]

            # Patrón de mitades por mes → cadencia.
            months: dict[tuple[int, int], set[str]] = {}
            days: list[int] = []
            quincenas: set[str] = set()
            for _id, oa, _amt, _cat, _pm in rows:
                d = datetime.fromtimestamp(oa / 1000, tz=timezone.utc).date()
                days.append(d.day)
                months.setdefault((d.year, d.month), set()).add(
                    "F" if d.day <= 15 else "S"
                )
            both = sum(1 for hs in months.values() if len(hs) == 2)
            both_ratio = both / len(months) if months else 0.0
            cadence = ("QUINCENAL_EVERY" if both_ratio >= 0.5
                       else "MONTHLY_SPECIFIC_HALF")
            days.sort()
            median_day = days[len(days) // 2]

            # Distintas quincenas → confidence.
            nq = len(set(
                cur.execute(
                    "SELECT quincena_id FROM expense WHERE concept=? AND status='POSTED'",
                    (concept,),
                ).fetchall()
            ))
            confidence = min(1.0, nq / 12.0)

            # Gasto más reciente → categoría, wallet, atribuciones default.
            recent = rows[-1]
            recent_id, _, _, category_id, payment_method_id = recent
            beneficiary_ids = [
                r[0] for r in cur.execute(
                    """SELECT member_id FROM expense_attribution
                       WHERE expense_id=? AND role='BENEFICIARY'
                       ORDER BY share_bps DESC""",
                    (recent_id,),
                ).fetchall()
            ]
            payer_rows = cur.execute(
                """SELECT member_id, share_bps FROM expense_attribution
                   WHERE expense_id=? AND role='PAYER'""",
                (recent_id,),
            ).fetchall()
            payer_split = {mid: bps for mid, bps in payer_rows}

            cadence_detail = json.dumps({"day_of_month": median_day})
            next_date = self._next_expected_date(cadence, median_day, TODAY)
            learned = json.dumps([r[0] for r in rows[:5]])
            tpl_id = did(f"rectpl:{_normalize(concept)}")

            cur.execute(
                """INSERT INTO recurrence_template
                       (id, household_id, concept, category_id,
                        default_amount_mxn, default_payment_method_id,
                        cadence, cadence_detail, next_expected_date,
                        default_beneficiary_ids, default_payer_split,
                        is_active, confidence_score, learned_from_expense_ids)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)""",
                (
                    tpl_id, HOUSEHOLD_ID, concept, category_id,
                    median_amount, payment_method_id,
                    cadence, cadence_detail, next_date,
                    json.dumps(beneficiary_ids),
                    json.dumps(payer_split),
                    confidence, learned,
                ),
            )
            created += 1
        conn.commit()
        LOG.info("Plantillas de recurrencia sembradas: %d", created)

    # ── paso 3.9: planes MSI reales (installment_plan) ──────────────────────

    def _seed_installment_plans(self, conn: sqlite3.Connection) -> None:
        """
        Siembra `installment_plan` para las series a meses reales documentadas
        en el concepto del Excel:
          * "Mercado Libre MSI" → 1 plan de 12 pagos (LOANS.MERCADO_LIBRE).
          * "Buró de Crédito"   → 1 plan de 3 pagos (LOANS.BURO).
        Deriva monto de cuota (mediana), principal (cuota × total),
        current_installment (pagos POSTED ≤ hoy), status (ACTIVE/PAID),
        start_date (primer pago) y wallet del propio gasto.

        NO se siembran Walmart/Coppel/Sears/Liverpool/Banamex como
        installment_plan: son crédito REVOLVENTE (saldo rotativo, no una serie
        cerrada de N cuotas), y ya viven como plantillas de recurrencia.
        """
        cur = conn.cursor()
        plans = (
            # (concept, key, display_name, total, category_code)
            ("Mercado Libre MSI", "mercado_libre",
             "Mercado Libre 12 MSI", 12, "LOANS.MERCADO_LIBRE"),
            ("Buró de Crédito", "buro",
             "Buró de Crédito 3 pagos", 3, "LOANS.BURO"),
        )
        created = 0
        for concept, key, display_name, total, cat_code in plans:
            rows = cur.execute(
                """SELECT amount_mxn, occurred_at, payment_method_id
                   FROM expense
                   WHERE concept = ? AND status = 'POSTED'
                   ORDER BY occurred_at""",
                (concept,),
            ).fetchall()
            if not rows:
                LOG.warning("Plan MSI omitido: sin gastos para %r", concept)
                continue

            amounts = sorted(r[0] for r in rows)
            installment_amount = amounts[len(amounts) // 2]  # mediana
            principal = round(installment_amount * total, 2)
            current = min(len(rows), total)  # pagos ya ocurridos (todos POSTED)
            status = "PAID" if current >= total else "ACTIVE"
            start_date = datetime.fromtimestamp(
                rows[0][1] / 1000, tz=timezone.utc
            ).date().isoformat()
            payment_method_id = rows[-1][2]
            category_id = CATEGORIES_BY_CODE[cat_code].id

            cur.execute(
                """INSERT INTO installment_plan
                       (id, household_id, display_name, creditor_member_id,
                        payment_method_id, principal_mxn, total_installments,
                        installment_amount_mxn, interest_rate_apr, start_date,
                        current_installment, status, category_id)
                   VALUES (?, ?, ?, NULL, ?, ?, ?, ?, 0.0, ?, ?, ?, ?)""",
                (
                    did(f"installment:{key}"), HOUSEHOLD_ID, display_name,
                    payment_method_id, principal, total,
                    installment_amount, start_date, current, status,
                    category_id,
                ),
            )
            created += 1
        conn.commit()
        LOG.info("Planes MSI sembrados: %d", created)

    def _finalize(self, conn: sqlite3.Connection) -> None:
        # Meta "Ahorro Empresa anual": current_mxn = suma de lo YA apartado
        # (celda E9 de cada hoja con período iniciado), acumulado en 3.6.
        # El target sigue siendo el seed de 60k; solo se actualiza el avance.
        conn.execute(
            "UPDATE savings_goal SET current_mxn = ? WHERE id = ?",
            (self._ahorro_empresa_mxn, did("savings_goal:ahorro_empresa")),
        )
        conn.commit()
        LOG.info("Ahorro Empresa acumulado a la fecha: $%.2f",
                 self._ahorro_empresa_mxn)

        # Reactiva FK para futuras sesiones (Room lo hace al abrir).
        conn.execute("PRAGMA foreign_keys = ON")
        # user_version = 1 es OBLIGATORIO: el asset declara schema v1. Si
        # queda en 0, el SQLiteOpenHelper de Android llama onCreate (no
        # onUpgrade) en instalación fresca y createAllTables de Room intenta
        # crear índices del schema actual sobre tablas del asset que no tienen
        # esas columnas → crash al primer arranque (footgun documentado en
        # CLAUDE.md). Debe fijarse ANTES del commit/cierre final.
        conn.execute("PRAGMA user_version = 1")
        # VACUUM compacta el archivo final (preserva user_version).
        conn.execute("VACUUM")
        conn.commit()


# ══════════════════════════════════════════════════════════════════════════════
#  BLOQUE 8 · CLI
# ══════════════════════════════════════════════════════════════════════════════

def main(argv: Optional[Iterable[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="ETL Excel quincenal -> Room SQLite (mx.budget).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--excel", type=Path,
                        default=Path("Copy of presupuesto 2.5.xlsx"),
                        help="Ruta al archivo Excel de origen.")
    parser.add_argument("--output", type=Path,
                        default=Path("budget_database.db"),
                        help="Ruta del SQLite de salida.")
    parser.add_argument("--rules", type=Path,
                        default=None,
                        help="Ruta al archivo attribution_rules.json (default: junto al script).")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="Activa logging DEBUG (incluye filas descartadas).")
    args = parser.parse_args(list(argv) if argv is not None else None)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    if not args.excel.exists():
        LOG.error("Archivo Excel no encontrado: %s", args.excel)
        return 2

    stats = EtlPipeline(args.excel, args.output, args.rules).run()

    # Reporte final al stdout.
    print()
    print("=" * 70)
    print(f"  Base generada: {args.output}")
    print("-" * 70)
    print(f"  Hojas procesadas:      {stats.sheets_parsed}")
    print(f"  Quincenas creadas:     {stats.quincenas_written}")
    print(f"  Ingresos creados:      {stats.incomes_written}")
    print(f"  Gastos creados:        {stats.expenses_written}")
    print(f"  Atribuciones creadas:  {stats.attributions_written}")
    print(f"  Reglas aplicadas:      {len(_ATTRIBUTION_RULES)}")
    if stats.skipped_sheets:
        print(f"  Hojas saltadas:        {stats.skipped_sheets}")
    print("=" * 70)
    return 0


if __name__ == "__main__":
    sys.exit(main())
